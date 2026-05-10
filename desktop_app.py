from __future__ import annotations

import json
import math
import os
import subprocess
import sys
import tkinter as tk
import ctypes
import re
from ctypes import wintypes
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Callable


APP_NAME = "Kadastro Harç Hesaplama"
YEAR = datetime.now().year
BASE_TARIFF_YEAR = 2026
MIN_KONTROLLUK = 2660.0
REV_FOOTER_TEXT = "B091TKG0010000.FR.232 Rev.No/Tarih:00/02.11.2009"
DEFAULT_LOGO_PATH = Path("C:/Users/Asus/OneDrive/Desktop/Tapu ve Kadastro Yeni Logo.jpg")
DEFAULT_ICON_PATH = Path("icon.png")
EXCEL_HEADERS = ["Bölge", "İl", "İlçe", "Tapu YK", "Kadastro YK"]
RATE_EXCEL_HEADERS = ["Kod", "Açıklama", "Tutar"]
CONTROL_RATE_DEFS: list[tuple[str, str, float]] = [
    ("min_kontrolluk", "Asgari kontrollük ücreti", 2660.0),
    ("mahkeme_infazi", "Mahkeme Kararları İnfazı", 3108.0),
    ("parselasyon_ha_1", "Parselasyon 0-100.000 m² (ha)", 3177.0),
    ("parselasyon_ha_2", "Parselasyon 100.000-500.000 m² (ha)", 2678.0),
    ("parselasyon_ha_3", "Parselasyon 500.000-1.000.000 m² (ha)", 2347.0),
    ("parselasyon_ha_4", "Parselasyon 1.000.000-1.500.000 m² (ha)", 1354.0),
    ("parselasyon_ha_5", "Parselasyon 1.500.000+ m² (ha)", 338.0),
    ("imar_deg_maktu_3000", "Değişiklik (imar) ilk dilim maktu", 2106.0),
    ("imar_deg_1000_5000", "Değişiklik (imar) 3.000-5.000 (1.000 m²)", 665.0),
    ("imar_deg_1000_10000", "Değişiklik (imar) 5.000-10.000 (1.000 m²)", 498.0),
    ("imar_deg_ha_100000", "Değişiklik (imar) 10.000-100.000 (ha)", 420.0),
    ("imar_deg_ha_2000000", "Değişiklik (imar) 100.000-2.000.000 (ha)", 338.0),
    ("degisiklik_parsel_ilave", "Değişiklik ilave parsel birim", 461.0),
    ("kad_deg_maktu_20000", "Değişiklik (kadastro) ilk dilim maktu", 1772.0),
    ("kad_deg_ha_100000", "Değişiklik (kadastro) 20.000-100.000 (ha)", 240.0),
    ("kad_deg_ha_4000000", "Değişiklik (kadastro) 100.000-4.000.000 (ha)", 143.0),
    ("kam_serit_kontrol", "Kamulaştırma şeritvari kontrol (km)", 4028.0),
    ("kam_serit_docs", "Kamulaştırma şeritvari bilgi-belge (km)", 892.0),
    ("kam_hektar_kontrol", "Kamulaştırma hektar kontrol (ha)", 892.0),
    ("kam_hektar_docs", "Kamulaştırma hektar bilgi-belge (ha)", 338.0),
    ("mulkiyet_parsel", "Mülkiyet raporu parsel başı", 235.0),
    ("mulkiyet_hektar", "Mülkiyet raporu hektar başı", 166.0),
    ("mulkiyet_km", "Mülkiyet raporu km başı", 427.0),
    ("plan_1000", "Plan/Kroki 0-1.000 m²", 2443.0),
    ("plan_3000", "Plan/Kroki 1.000-3.000 m²", 3676.0),
    ("plan_5000", "Plan/Kroki 3.000-5.000 m²", 7355.0),
    ("plan_10000", "Plan/Kroki 5.000-10.000 m²", 9794.0),
    ("plan_20000", "Plan/Kroki 10.000-20.000 m²", 13474.0),
    ("plan_50000", "Plan/Kroki 20.000-50.000 m²", 17590.0),
    ("plan_100000", "Plan/Kroki 50.000-100.000 m²", 21389.0),
    ("plan_200000", "Plan/Kroki 100.000-200.000 m²", 24508.0),
    ("plan_500000", "Plan/Kroki 200.000-500.000 m²", 29389.0),
    ("plan_extra_100000", "Plan/Kroki 500.000+ ilave 100.000 m²", 2008.0),
    ("cins_1000", "Cins değişikliği 0-1.000 m²", 4018.0),
    ("cins_3000", "Cins değişikliği 1.000-3.000 m²", 5550.0),
    ("cins_extra_1000", "Cins değişikliği 3.000+ ilave 1.000 m²", 509.0),
    ("cins_tarimsal_tavan", "Cins değişikliği tarımsal tavan", 12762.0),
    ("cins_yapisiz_parsel", "Cins yapıdan yapısıza parsel başı", 1033.0),
    ("cins_yerinde_gosterim", "Parsel yerinde gösterim (2.5.1)", 1003.0),
    ("cins_kat_ilk", "Cins kat ilavesi ilk yapı", 2170.0),
    ("cins_kat_ilave", "Cins kat ilavesi ilave yapı", 1023.0),
    ("cins_yeni_bina", "Cins yeni yapılan bina", 1609.0),
    ("irtifak_parsel", "İrtifak parsel başı", 2008.0),
    ("aplikasyon_kimlik", "Aplikasyon kimlik güncelleme", 2215.0),
    ("mera_gm_aplikasyon", "Mera GM aplikasyon (ha)", 1131.0),
    ("mera_komisyon_aplikasyon", "Mera komisyon aplikasyon (ha)", 272.0),
    ("kadastral_yol_10", "Kadastral yol ilk 10 nokta", 4971.0),
    ("kadastral_yol_ilave", "Kadastral yol ilave nokta", 235.0),
]
CONTROL_RATE_DEFAULTS: dict[str, float] = {k: v for k, _l, v in CONTROL_RATE_DEFS}
try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_IMPORT_ERROR: Exception | None = None
except Exception as exc:  # pragma: no cover - depends on runtime packaging
    Workbook = None  # type: ignore[assignment]
    load_workbook = None  # type: ignore[assignment]
    OPENPYXL_IMPORT_ERROR = exc


def resource_path(relative: str) -> Path:
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return base / relative


def user_data_dir() -> Path:
    root = os.environ.get("APPDATA") or str(Path.home() / "AppData" / "Roaming")
    path = Path(root) / "KadastroHarcDesktop"
    path.mkdir(parents=True, exist_ok=True)
    return path


SETTINGS_FILE = user_data_dir() / "ayarlar.json"
TR_UPPER_TRANS = str.maketrans({"i": "İ", "ı": "I"})
TR_LOWER_TRANS = str.maketrans({"I": "ı", "İ": "i"})


def tr_money(value: float) -> str:
    text = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{text} TL"


def tr_upper(value: str) -> str:
    return str(value or "").translate(TR_UPPER_TRANS).upper()


def tr_lower(value: str) -> str:
    return str(value or "").translate(TR_LOWER_TRANS).lower()


def tr_capitalize(value: str) -> str:
    if not value:
        return ""
    return tr_upper(value[0]) + tr_lower(value[1:])


def tr_title(value: str) -> str:
    out: list[str] = []
    current: list[str] = []
    separators = set(" -/(),.")
    for ch in str(value or ""):
        if ch in separators:
            if current:
                out.append(tr_capitalize("".join(current)))
                current = []
            out.append(ch)
        else:
            current.append(ch)
    if current:
        out.append(tr_capitalize("".join(current)))
    return "".join(out)


def normalize_key(value: str) -> str:
    return tr_upper(str(value or "").strip())


def mudurluk_name_from_il(il: str) -> str:
    clean = normalize_key(il)
    if not clean:
        return "Kadastro Müdürlüğü"
    return f"{tr_title(clean)} Kadastro Müdürlüğü"


def parse_decimal_input(value: str | float | int | None, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip().replace(" ", "")
    if not raw:
        return default
    if "," in raw and "." in raw:
        if raw.rfind(",") > raw.rfind("."):
            raw = raw.replace(".", "").replace(",", ".")
        else:
            raw = raw.replace(",", "")
    elif "," in raw:
        raw = raw.replace(",", ".")
    return float(raw)


def canonical_header(value: str) -> str:
    text = normalize_key(value)
    for token in [" ", "_", "-", "/", "(", ")"]:
        text = text.replace(token, "")
    return text


def pick_rate(rate_map: dict[str, float] | None, key: str, default: float) -> float:
    if not isinstance(rate_map, dict):
        return float(default)
    value = rate_map.get(key, default)
    try:
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def resolve_display_name() -> str:
    # 1) Domain/local display name from Windows API (best effort).
    try:
        name_display = 3  # EXTENDED_NAME_FORMAT.NameDisplay
        secur32 = ctypes.WinDLL("secur32", use_last_error=True)
        get_name = secur32.GetUserNameExW
        get_name.argtypes = [wintypes.ULONG, wintypes.LPWSTR, ctypes.POINTER(wintypes.ULONG)]
        get_name.restype = wintypes.BOOL
        size = wintypes.ULONG(0)
        get_name(name_display, None, ctypes.byref(size))
        if size.value > 1:
            buf = ctypes.create_unicode_buffer(size.value)
            if get_name(name_display, buf, ctypes.byref(size)) and buf.value.strip():
                return buf.value.strip()
    except Exception:
        pass

    # 2) Try AD account display name through PowerShell.
    try:
        cmd = (
            "Add-Type -AssemblyName System.DirectoryServices.AccountManagement; "
            "[System.DirectoryServices.AccountManagement.UserPrincipal]::Current.DisplayName"
        )
        result = subprocess.run(
            ["powershell", "-NoProfile", "-Command", cmd],
            capture_output=True,
            text=True,
            timeout=4,
            check=False,
        )
        value = result.stdout.strip()
        if value:
            return value
    except Exception:
        pass

    # 3) Local user full name (when available).
    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-Command", "Get-LocalUser -Name $env:USERNAME | Select-Object -ExpandProperty FullName"],
            capture_output=True,
            text=True,
            timeout=4,
            check=False,
        )
        value = result.stdout.strip()
        if value:
            return value
    except Exception:
        pass

    # 4) Fallback: username.
    return os.environ.get("USERNAME", "Kullanıcı")


def round_lira(value: float) -> float:
    return float(math.floor(float(value) + 0.5))


def normalize_area(value: float) -> float:
    return max(float(value or 0), 0.0)


def billed_hectare(value: float, minimum: int = 1, maximum: int | None = None) -> float:
    raw = max(float(value or 0), float(minimum))
    stepped = float(math.ceil(raw))
    if maximum is not None:
        stepped = min(stepped, float(maximum))
    return stepped


def stepped_area(area_m2: float, brackets: list[tuple[float | None, float, str]]) -> tuple[float, list[str]]:
    area = normalize_area(area_m2)
    previous = 0.0
    total = 0.0
    details: list[str] = []
    for upper, unit_price, unit in brackets:
        limit = area if upper is None else min(area, upper)
        amount = max(limit - previous, 0.0)
        if amount <= 0:
            previous = upper or previous
            continue
        divisor = 10000.0 if unit == "ha" else 1000.0
        line_total = amount / divisor * unit_price
        total += line_total
        label = "ha" if unit == "ha" else "1.000 m2"
        details.append(f"{amount:,.2f} m2 / {label} x {tr_money(unit_price)} = {tr_money(line_total)}")
        previous = limit
        if upper is not None and area <= upper:
            break
    return total, details


def parselasyon_base_fee(area_m2: float, rate_map: dict[str, float] | None = None) -> tuple[float, list[str]]:
    area = normalize_area(area_m2)
    p1 = pick_rate(rate_map, "parselasyon_ha_1", 3177.0)
    p2 = pick_rate(rate_map, "parselasyon_ha_2", 2678.0)
    p3 = pick_rate(rate_map, "parselasyon_ha_3", 2347.0)
    p4 = pick_rate(rate_map, "parselasyon_ha_4", 1354.0)
    p5 = pick_rate(rate_map, "parselasyon_ha_5", 338.0)
    base, details = stepped_area(
        area,
        [
            (100000, p1, "ha"),
            (500000, p2, "ha"),
            (1000000, p3, "ha"),
            (1500000, p4, "ha"),
            (None, p5, "ha"),
        ],
    )
    return base, details


def apply_parselasyon_special(
    base: float,
    details: list[str],
    geriye_donus: bool = False,
    yeni_uygulama: bool = False,
) -> tuple[float, list[str]]:
    out = list(details)
    if geriye_donus and not yeni_uygulama:
        out.append("Geriye dönüş: hesaplanan kontrollük bedelinin %25'i tahsil edilir.")
        base *= 0.25
    elif geriye_donus and yeni_uygulama:
        out.append("Geriye dönüş + yeni uygulama: kontrollük bedeline ayrıca %25 eklenir.")
        base *= 1.25
    return base, out


def parselasyon_fee(
    area_m2: float,
    yk: float,
    geriye_donus: bool = False,
    yeni_uygulama: bool = False,
    rate_map: dict[str, float] | None = None,
) -> tuple[float, list[str]]:
    base, details = parselasyon_base_fee(area_m2, rate_map)
    base, details = apply_parselasyon_special(base, details, geriye_donus, yeni_uygulama)
    details.append(f"Yöresel katsayı: {yk:g}")
    return round_lira(base * yk), details


def toplulastirma_ilave_basamak_fee(
    current_area_m2: float,
    previous_area_m2: float,
    yk: float,
    geriye_donus: bool = False,
    yeni_uygulama: bool = False,
    rate_map: dict[str, float] | None = None,
) -> tuple[float, list[str]]:
    current_area = normalize_area(current_area_m2)
    previous_area = normalize_area(previous_area_m2)
    total_area = previous_area + current_area

    total_base, total_details = parselasyon_base_fee(total_area, rate_map)
    previous_base, previous_details = parselasyon_base_fee(previous_area, rate_map)
    difference_base = max(total_base - previous_base, 0.0)

    details = [
        "Arazi toplulaştırma hesabı, daha önce bedeli tahsil edilen basamağa ilave edilerek yapılmıştır.",
        f"Daha önce aynı sözleşme kapsamında dikkate alınan alan: {previous_area:,.2f} m²",
        f"Bu kontrole gelen alan: {current_area:,.2f} m²",
        f"Toplam kümülatif alan: {total_area:,.2f} m²",
        "",
        "Toplam kümülatif alanın kademeli hesabı:",
        *[f"  - {line}" for line in total_details],
        f"Toplam kümülatif baz bedel: {tr_money(total_base)}",
        "",
        "Daha önce dikkate alınan alanın kademeli hesabı:",
    ]

    if previous_area > 0:
        details.extend([f"  - {line}" for line in previous_details])
        details.append(f"Daha önceki alanın baz bedeli: {tr_money(previous_base)}")
    else:
        details.append("  - Önceki alan yok.")
        details.append(f"Daha önceki alanın baz bedeli: {tr_money(0)}")

    details.extend(["", f"Bu tahakkukta hesaplanan ilave basamak bedeli: {tr_money(difference_base)}"])

    adjusted_base, special_details = apply_parselasyon_special(
        difference_base,
        [],
        geriye_donus,
        yeni_uygulama,
    )
    details.extend(special_details)
    details.append(f"Yöresel katsayı: {yk:g}")
    return round_lira(adjusted_base * yk), details


def imar_degisiklik_fee(
    area_m2: float,
    yk: float,
    parsel_count: int,
    mera_25: bool,
    rate_map: dict[str, float] | None = None,
) -> tuple[float, list[str]]:
    area = normalize_area(area_m2)
    imar_maktu = pick_rate(rate_map, "imar_deg_maktu_3000", 2106.0)
    imar_1000_5000 = pick_rate(rate_map, "imar_deg_1000_5000", 665.0)
    imar_1000_10000 = pick_rate(rate_map, "imar_deg_1000_10000", 498.0)
    imar_ha_100000 = pick_rate(rate_map, "imar_deg_ha_100000", 420.0)
    imar_ha_2000000 = pick_rate(rate_map, "imar_deg_ha_2000000", 338.0)
    ilave_parsel = pick_rate(rate_map, "degisiklik_parsel_ilave", 461.0)
    if area <= 3000:
        base = imar_maktu
        details = [f"İlk dilim maktu = {tr_money(base)}"]
    else:
        base, details = stepped_area(
            area,
            [
                (3000, imar_maktu / 3.0, "1000"),
                (5000, imar_1000_5000, "1000"),
                (10000, imar_1000_10000, "1000"),
                (100000, imar_ha_100000, "ha"),
                (2000000, imar_ha_2000000, "ha"),
            ],
        )
    extra = max(int(parsel_count or 0) - 2, 0) * ilave_parsel
    if extra:
        details.append(f"İkiden fazla parsel ilavesi = {tr_money(extra)}")
        base += extra
    if mera_25:
        details.append("Mera Kanunu özel durumu: %25 tahsil")
        base *= 0.25
    return base * yk, details


def kadastro_degisiklik_fee(
    area_m2: float,
    yk: float,
    parsel_count: int,
    mera_25: bool,
    rate_map: dict[str, float] | None = None,
) -> tuple[float, list[str]]:
    area = normalize_area(area_m2)
    kad_maktu = pick_rate(rate_map, "kad_deg_maktu_20000", 1772.0)
    kad_ha_100000 = pick_rate(rate_map, "kad_deg_ha_100000", 240.0)
    kad_ha_4000000 = pick_rate(rate_map, "kad_deg_ha_4000000", 143.0)
    ilave_parsel = pick_rate(rate_map, "degisiklik_parsel_ilave", 461.0)
    if area <= 20000:
        base = kad_maktu
        details = [f"İlk dilim maktu = {tr_money(base)}"]
    else:
        base, details = stepped_area(
            area,
            [
                (20000, kad_maktu / 2.0, "ha"),
                (100000, kad_ha_100000, "ha"),
                (4000000, kad_ha_4000000, "ha"),
            ],
        )
    extra = max(int(parsel_count or 0) - 2, 0) * ilave_parsel
    if extra:
        details.append(f"İkiden fazla parsel ilavesi = {tr_money(extra)}")
        base += extra
    if mera_25:
        details.append("Mera Kanunu özel durumu: %25 tahsil")
        base *= 0.25
    return base * yk, details


def kamulastirma_serit_fee(km: float, include_docs: bool, rate_map: dict[str, float] | None = None) -> tuple[float, list[str]]:
    effective_km = max(float(km or 0), 1.0)
    serit_kontrol = pick_rate(rate_map, "kam_serit_kontrol", 4028.0)
    serit_docs = pick_rate(rate_map, "kam_serit_docs", 892.0)
    control = effective_km * serit_kontrol
    total = control
    details = [f"Şeritvari kontrol: {effective_km:g} km x {tr_money(serit_kontrol)} = {tr_money(control)}"]
    if include_docs:
        docs = effective_km * serit_docs
        total += docs
        details.append(f"Harita bilgi ve belge: {effective_km:g} km x {tr_money(serit_docs)} = {tr_money(docs)}")
    return total, details


def kamulastirma_hektar_fee(
    hektar: float,
    yk: float,
    include_docs: bool,
    mera_25: bool,
    rate_map: dict[str, float] | None = None,
) -> tuple[float, list[str]]:
    ha = billed_hectare(hektar, minimum=1, maximum=100)
    hektar_kontrol = pick_rate(rate_map, "kam_hektar_kontrol", 892.0)
    hektar_docs = pick_rate(rate_map, "kam_hektar_docs", 338.0)
    control = ha * hektar_kontrol
    total = control
    details = [f"Hektar bazlı kontrol: {ha:g} ha x {tr_money(hektar_kontrol)} = {tr_money(control)}"]
    if include_docs:
        docs = ha * hektar_docs
        total += docs
        details.append(f"Harita bilgi ve belge: {ha:g} ha x {tr_money(hektar_docs)} = {tr_money(docs)}")
    if mera_25:
        details.append("Mera Kanunu özel durumu: %25 tahsil")
        total *= 0.25
    return total * yk, details


def mulkiyet_raporu_fee(
    report_type: str,
    hektar: float,
    km: float,
    parsel_count: int,
    rate_map: dict[str, float] | None = None,
) -> tuple[float, list[str]]:
    mul_parsel = pick_rate(rate_map, "mulkiyet_parsel", 235.0)
    mul_hektar = pick_rate(rate_map, "mulkiyet_hektar", 166.0)
    mul_km = pick_rate(rate_map, "mulkiyet_km", 427.0)
    if report_type.startswith("İmar ayırma"):
        adet = max(int(parsel_count or 0), 1)
        total = adet * mul_parsel
        return total, [f"Parsel başı mülkiyet raporu: {adet} x {tr_money(mul_parsel)} = {tr_money(total)}"]
    if report_type.startswith("Parselasyon"):
        raw_ha = max(float(hektar or 0), 1.0)
        hektar = billed_hectare(raw_ha, minimum=1)
        total = hektar * mul_hektar
        details = [f"Hektar başı mülkiyet raporu: {hektar:g} ha x {tr_money(mul_hektar)} = {tr_money(total)}"]
        if hektar != raw_ha:
            details.append(f"Küsuratlı hektar üst tam sayıya tamamlandı: {raw_ha:g} ha -> {hektar:g} ha")
        return total, details
    eff_km = max(float(km or 0), 1.0)
    total = eff_km * mul_km
    return total, [f"Km başı mülkiyet raporu: {eff_km:g} x {tr_money(mul_km)} = {tr_money(total)}"]


def irtifak_fee(
    parsel_count: int,
    irtifak_type: str,
    yk: float,
    hektar: float = 0.0,
    km: float = 0.0,
    include_docs: bool = False,
    corner_qty: int = 0,
    poligon_qty: int = 0,
    olcu_kroki_qty: int = 0,
    aplikasyon_kroki_qty: int = 0,
    rate_map: dict[str, float] | None = None,
) -> tuple[float, list[str], bool]:
    count = max(int(parsel_count or 0), 1)
    details: list[str] = []
    uses_yk = False
    irtifak_parsel = pick_rate(rate_map, "irtifak_parsel", 2008.0)
    hektar_kontrol = pick_rate(rate_map, "kam_hektar_kontrol", 892.0)
    hektar_docs = pick_rate(rate_map, "kam_hektar_docs", 338.0)
    serit_kontrol = pick_rate(rate_map, "kam_serit_kontrol", 4028.0)
    serit_docs = pick_rate(rate_map, "kam_serit_docs", 892.0)

    if irtifak_type.startswith("Standart"):
        raw = count * irtifak_parsel
        details.append(f"2.4 İrtifak tesisi/terkini: {count} parsel x {tr_money(irtifak_parsel)} = {tr_money(raw)}")
        details.append("Yüzölçümü ve yöresel katsayı uygulanmaz.")
        return raw, details, uses_yk

    if "2.11.5.2.4" in irtifak_type:
        raw = count * irtifak_parsel
        details.append(f"2.11.5.2.4 gereği 2.4 uygulanır: {count} parsel x {tr_money(irtifak_parsel)} = {tr_money(raw)}")
        details.append("Kontrollük hektar bedeli uygulanmaz; kullanılan 4. bölüm bilgi-belge bedelleri eklenir.")
        if include_docs:
            c_qty = max(int(corner_qty or 0), 0)
            p_qty = max(int(poligon_qty or 0), 0)
            o_qty = max(int(olcu_kroki_qty or 0), 0)
            a_qty = max(int(aplikasyon_kroki_qty or 0), 0)

            corner_total, corner_label = corner_coordinate_fee(c_qty)
            poligon_total = p_qty * hektar_docs
            olcu_total = o_qty * 307.0
            aplikasyon_total = a_qty * 143.0
            docs_total = corner_total + poligon_total + olcu_total + aplikasyon_total
            raw += docs_total

            details.append("4. bölüm harita bilgi-belge bedelleri:")
            if c_qty:
                details.append(f"Parsel köşe koordinatı: {c_qty} adet | {corner_label} = {tr_money(corner_total)}")
            if p_qty:
                details.append(f"Poligon koordinatı: {p_qty} x {tr_money(hektar_docs)} = {tr_money(poligon_total)}")
            if o_qty:
                details.append(f"Ölçü krokisi: {o_qty} x {tr_money(307)} = {tr_money(olcu_total)}")
            if a_qty:
                details.append(f"Aplikasyon krokisi: {a_qty} x {tr_money(143)} = {tr_money(aplikasyon_total)}")
            details.append(f"Bilgi-belge toplamı: {tr_money(docs_total)}")
        else:
            details.append("Harita bilgi-belge bedeli eklenmedi.")
        return raw, details, uses_yk

    if "hektar bazlı kontrol" in tr_lower(irtifak_type):
        uses_yk = True
        raw_ha = max(float(hektar or 0), 1.0)
        control_ha = min(raw_ha, 100.0)
        control = control_ha * hektar_kontrol
        docs = raw_ha * hektar_docs if include_docs else 0.0
        raw = (control + docs) * yk
        details.append(f"Hektar bazlı kontrol: {control_ha:g} ha x {tr_money(hektar_kontrol)} = {tr_money(control)}")
        if raw_ha > 100.0:
            details.append("Kontrollük kısmında 100 ha üst sınırı uygulandı.")
        if include_docs:
            details.append(f"Hektar bazlı harita bilgi-belge: {raw_ha:g} ha x {tr_money(hektar_docs)} = {tr_money(docs)}")
        details.append(f"Yöresel katsayı uygulandı: {yk:g}")
        return raw, details, uses_yk

    if "şeritvari" in tr_lower(irtifak_type):
        eff_km = max(float(km or 0), 1.0)
        control = eff_km * serit_kontrol
        docs = eff_km * serit_docs if include_docs else 0.0
        raw = control + docs
        details.append(f"Şeritvari kontrol: {eff_km:g} km x {tr_money(serit_kontrol)} = {tr_money(control)}")
        if include_docs:
            details.append(f"Şeritvari harita bilgi-belge: {eff_km:g} km x {tr_money(serit_docs)} = {tr_money(docs)}")
        details.append("Yöresel katsayı uygulanmaz.")
        return raw, details, uses_yk

    raw = count * irtifak_parsel
    details.append(f"Varsayılan 2.4 hesabı: {count} x {tr_money(irtifak_parsel)} = {tr_money(raw)}")
    return raw, details, uses_yk


def kadastral_yol_fee(point_count: int, rate_map: dict[str, float] | None = None) -> tuple[float, list[str]]:
    count = max(int(point_count or 0), 1)
    first_10 = pick_rate(rate_map, "kadastral_yol_10", 4971.0)
    extra_unit = pick_rate(rate_map, "kadastral_yol_ilave", 235.0)
    if count <= 10:
        total = first_10
        details = [
            f"Kadastral yol sınırları: {count} nokta, 10 noktaya kadar maktu {tr_money(first_10)}",
            "Yöresel katsayı uygulanmaz.",
        ]
    else:
        extra_count = count - 10
        extra = extra_count * extra_unit
        total = first_10 + extra
        details = [
            f"Kadastral yol sınırları: İlk 10 nokta = {tr_money(first_10)}",
            f"İlave {extra_count} nokta x {tr_money(extra_unit)} = {tr_money(extra)}",
            "Yöresel katsayı uygulanmaz.",
        ]
    return total, details


def aplikasyon_fee(
    aplikasyon_type: str,
    area_m2: float,
    yk: float,
    parsel_turu: str,
    parsel_count: int = 1,
    contract_amount: float = 0.0,
    hectare: float = 0.0,
    point_count: int = 0,
    rate_map: dict[str, float] | None = None,
) -> tuple[float, list[str], bool]:
    details: list[str] = []
    uses_yk = False

    if aplikasyon_type.startswith("Normal aplikasyon"):
        raw, details = plan_kroki_fee(area_m2, yk, parsel_turu, rate_map)
        details.insert(0, "2.1 Aplikasyon işlemi yüzölçümü tarifesi uygulandı.")
        uses_yk = True
        return raw, details, uses_yk

    if aplikasyon_type.startswith("Kimlik bilgisi"):
        count = max(int(parsel_count or 0), 1)
        kimlik_birim = pick_rate(rate_map, "aplikasyon_kimlik", 2215.0)
        raw = kimlik_birim
        details.append(f"1.10.1 Kimlik bilgisi güncelleme için zeminde inceleme: {tr_money(kimlik_birim)}")
        if count > 1:
            details.append(
                f"{count} taşınmaz girildi. 1.10.2 ilave taşınmazlar için ilave gösterge gerektirir; "
                "ilave gösterge sabiti tanımlı olmadığı için yalnızca ilk taşınmaz hesaplandı."
            )
        details.append("Yöresel katsayı uygulanmaz.")
        return raw, details, uses_yk

    if aplikasyon_type.startswith("Mera ihale aplikasyon kontrolü"):
        amount = max(float(contract_amount or 0.0), 0.0)
        raw = amount * 0.10
        details.append(f"2.8.2 Mera aplikasyon kontrolü: sözleşme bedeli {tr_money(amount)} x %10 = {tr_money(raw)}")
        details.append("Yöresel katsayı uygulanmaz.")
        return raw, details, uses_yk

    if aplikasyon_type.startswith("Mera GM hizmet talebi"):
        uses_yk = True
        ha = billed_hectare(hectare, minimum=1)
        gm_unit = pick_rate(rate_map, "mera_gm_aplikasyon", 1131.0)
        base = ha * gm_unit
        raw = base * yk
        details.append(f"2.8.3.2 Mera aplikasyon yapımı: {ha:g} ha x {tr_money(gm_unit)} = {tr_money(base)}")
        details.append(f"Yöresel katsayı uygulandı: {yk:g}")
        return raw, details, uses_yk

    if aplikasyon_type.startswith("Mera Komisyonu aplikasyon kontrolü"):
        uses_yk = True
        ha = billed_hectare(hectare, minimum=1)
        kom_unit = pick_rate(rate_map, "mera_komisyon_aplikasyon", 272.0)
        base = ha * kom_unit
        raw = base * yk
        details.append(f"2.8.4.1 Mera Komisyonu aplikasyon kontrolü: {ha:g} ha x {tr_money(kom_unit)} = {tr_money(base)}")
        details.append(f"Yöresel katsayı uygulandı: {yk:g}")
        return raw, details, uses_yk

    if aplikasyon_type.startswith("Kadastral yol"):
        raw, details = kadastral_yol_fee(point_count, rate_map)
        return raw, details, uses_yk

    raw, details = plan_kroki_fee(area_m2, yk, parsel_turu, rate_map)
    details.insert(0, "Varsayılan olarak 2.1 normal aplikasyon tarifesi uygulandı.")
    uses_yk = True
    return raw, details, uses_yk


def plan_kroki_base(area: float, rate_map: dict[str, float] | None = None) -> tuple[float, str]:
    p1000 = pick_rate(rate_map, "plan_1000", 2443.0)
    p3000 = pick_rate(rate_map, "plan_3000", 3676.0)
    p5000 = pick_rate(rate_map, "plan_5000", 7355.0)
    p10000 = pick_rate(rate_map, "plan_10000", 9794.0)
    p20000 = pick_rate(rate_map, "plan_20000", 13474.0)
    p50000 = pick_rate(rate_map, "plan_50000", 17590.0)
    p100000 = pick_rate(rate_map, "plan_100000", 21389.0)
    p200000 = pick_rate(rate_map, "plan_200000", 24508.0)
    p500000 = pick_rate(rate_map, "plan_500000", 29389.0)
    pextra = pick_rate(rate_map, "plan_extra_100000", 2008.0)
    if area <= 0:
        return 0.0, "Yüzölçümü girilmedi"
    if area <= 1000:
        return p1000, "1 < A <= 1.000 m²"
    if area <= 3000:
        return p3000, "1.000 < A <= 3.000 m²"
    if area <= 5000:
        return p5000, "3.000 < A <= 5.000 m²"
    if area <= 10000:
        return p10000, "5.000 < A <= 10.000 m²"
    if area <= 20000:
        return p20000, "10.000 < A <= 20.000 m²"
    if area <= 50000:
        return p50000, "20.000 < A <= 50.000 m²"
    if area <= 100000:
        return p100000, "50.000 < A <= 100.000 m²"
    if area <= 200000:
        return p200000, "100.000 < A <= 200.000 m²"
    if area <= 500000:
        return p500000, "200.000 < A <= 500.000 m²"

    extra_blocks = math.ceil((area - 500000.0) / 100000.0)
    base = p500000 + (extra_blocks * pextra)
    return base, "A > 500.000 m²: 2.1.9 + her 100.000 m² için 2.008 TL"


def plan_kroki_fee(
    area_m2: float,
    yk: float,
    parsel_turu: str,
    rate_map: dict[str, float] | None = None,
) -> tuple[float, list[str]]:
    area = normalize_area(area_m2)
    if area <= 0:
        return 0.0, ["Yüzölçümü girilmedi."]
    base, detail = plan_kroki_base(area, rate_map)

    factor = 1.0
    factor_label = "İmar parseli (%100)"
    if "Kadastro" in parsel_turu:
        factor = 0.5
        factor_label = "Kadastro parseli (%50)"
    if "8/ğ" in parsel_turu or "Belediye Dışı" in parsel_turu:
        factor = 0.25
        factor_label = "8/ğ / Belediye dışı (%25)"

    total = base * factor * yk
    details = [
        f"{detail}: {tr_money(base)}",
        f"Parsel türü katsayısı: {factor_label}",
        f"Yöresel katsayı: {yk:g}",
    ]
    return total, details


def cins_degisikligi_fee(
    area_m2: float,
    yk: float,
    parsel_turu: str,
    rate_map: dict[str, float] | None = None,
) -> tuple[float, list[str]]:
    area = normalize_area(area_m2)
    c1000 = pick_rate(rate_map, "cins_1000", 4018.0)
    c3000 = pick_rate(rate_map, "cins_3000", 5550.0)
    cextra = pick_rate(rate_map, "cins_extra_1000", 509.0)
    if area <= 0:
        return 0.0, ["Yüzölçümü girilmedi."]
    if area <= 1000:
        base = c1000
        detail = "1 < A <= 1.000 m²"
    elif area <= 3000:
        base = c3000
        detail = "1.000 < A <= 3.000 m²"
    else:
        extra = ((area - 3000.0) / 1000.0) * cextra
        base = c1000 + c3000 + extra
        detail = "3.000 m² üstü ilave 509 TL/1.000 m²"
    factor = 1.0
    if "Kadastro" in parsel_turu:
        factor = 0.5
    total = base * factor * yk
    details = [
        f"{detail}: {tr_money(base)}",
        f"Parsel türü katsayısı: {'%50' if factor < 1 else '%100'}",
        f"Yöresel katsayı: {yk:g}",
    ]
    return total, details


def cins_degisikligi_detayli_fee(
    cins_type: str,
    area_m2: float,
    yk: float,
    parsel_turu: str,
    tarimsal_tavan: bool = False,
    fazla_yapi_sayisi: int = 0,
    parsel_sayisi: int = 1,
    yapi_sayisi: int = 1,
    arsa_pay_orani: float = 100.0,
    rate_map: dict[str, float] | None = None,
) -> tuple[float, list[str], bool]:
    details: list[str] = []
    uses_yk = True

    if cins_type.startswith("Yapısız"):
        raw, details = cins_degisikligi_fee(area_m2, yk, parsel_turu, rate_map)
        tarimsal_tavan_limit = pick_rate(rate_map, "cins_tarimsal_tavan", 12762.0)
        if tarimsal_tavan and raw > tarimsal_tavan_limit:
            details.append(f"Tarımsal amaçlı bina tavanı uygulandı: {tr_money(tarimsal_tavan_limit)}")
            raw = tarimsal_tavan_limit
        fazla = max(int(fazla_yapi_sayisi or 0), 0)
        if fazla:
            yapi_ilave = pick_rate(rate_map, "cins_yapisiz_parsel", 1033.0)
            extra = fazla * yapi_ilave
            raw += extra
            details.append(f"Fazla yapı ilavesi: {fazla} x {tr_money(yapi_ilave)} = {tr_money(extra)}")
        return raw, details, uses_yk

    if cins_type.startswith("Yapılı iken yapısız"):
        uses_yk = False
        count = max(int(parsel_sayisi or 0), 1)
        yapi_parsel_birim = pick_rate(rate_map, "cins_yapisiz_parsel", 1033.0)
        raw = count * yapi_parsel_birim
        details.append(f"Yapılı iken yapısız hâle gelme: {count} parsel x {tr_money(yapi_parsel_birim)} = {tr_money(raw)}")
        details.append("Yöresel katsayı uygulanmaz, maktu ücrettir.")
        return raw, details, uses_yk

    if cins_type.startswith("Yapı ile ilgisi olmayan"):
        uses_yk = False
        count = max(int(parsel_sayisi or 0), 1)
        yapi_parsel_birim = pick_rate(rate_map, "cins_yapisiz_parsel", 1033.0)
        yerinde_birim = pick_rate(rate_map, "cins_yerinde_gosterim", 1003.0)
        unit = yapi_parsel_birim + yerinde_birim
        raw = count * unit
        details.append(f"2.2.2 ücreti: {tr_money(yapi_parsel_birim)}")
        details.append(f"2.5.1 parselin yerinde gösterilmesi ücreti: {tr_money(yerinde_birim)}")
        details.append(f"Vasıf değişikliği: {count} parsel x {tr_money(unit)} = {tr_money(raw)}")
        details.append("Yöresel katsayı uygulanmaz, maktu ücrettir.")
        return raw, details, uses_yk

    if cins_type.startswith("Aynı yapı üzerinde kat ilavesi"):
        uses_yk = False
        kat_ilk = pick_rate(rate_map, "cins_kat_ilk", 2170.0)
        raw = kat_ilk
        details.append(f"Aynı yapı üzerinde kat ilavesi: {tr_money(kat_ilk)}")
        details.append("Yöresel katsayı uygulanmaz, maktu ücrettir.")
        return raw, details, uses_yk

    if cins_type.startswith("Birden fazla yapı varsa kat ilavesi"):
        uses_yk = False
        count = max(int(yapi_sayisi or 0), 1)
        kat_ilk = pick_rate(rate_map, "cins_kat_ilk", 2170.0)
        kat_ilave = pick_rate(rate_map, "cins_kat_ilave", 1023.0)
        raw = kat_ilk + max(count - 1, 0) * kat_ilave
        details.append(f"İlk yapı kat ilavesi: {tr_money(kat_ilk)}")
        if count > 1:
            details.append(f"İlave yapı kat ilavesi: {count - 1} x {tr_money(kat_ilave)} = {tr_money((count - 1) * kat_ilave)}")
        details.append("Yöresel katsayı uygulanmaz, maktu ücrettir.")
        return raw, details, uses_yk

    if cins_type.startswith("Aynı parselde sonradan"):
        uses_yk = False
        count = max(int(yapi_sayisi or 0), 1)
        yeni_bina = pick_rate(rate_map, "cins_yeni_bina", 1609.0)
        raw = count * yeni_bina
        details.append(f"Yeni yapılan bina: {count} x {tr_money(yeni_bina)} = {tr_money(raw)}")
        details.append("Yöresel katsayı uygulanmaz, maktu ücrettir.")
        return raw, details, uses_yk

    if cins_type.startswith("Yaygın kat mülkiyeti"):
        raw, details = cins_degisikligi_fee(area_m2, yk, parsel_turu, rate_map)
        oran = max(min(float(arsa_pay_orani or 0), 100.0), 0.0)
        partial = raw * oran / 100.0
        details.append(f"Yaygın kat mülkiyeti kısmi hesap: {tr_money(raw)} x %{oran:g} = {tr_money(partial)}")
        return partial, details, uses_yk

    raw, details = cins_degisikligi_fee(area_m2, yk, parsel_turu, rate_map)
    return raw, details, uses_yk


@dataclass(frozen=True)
class TechnicalItem:
    key: str
    label: str
    unit: str
    price: float


MAIN_TECH_ITEMS = [
    TechnicalItem("parcel_corner", "Parsel Köşe Noktası", "Nokta", 0.0),
    TechnicalItem("poligon_xy", "Poligon Koordinatı (x - y)", "Adet", 338.0),
    TechnicalItem("olcu_kroki", "Ölçü Krokisi", "Sayfa", 307.0),
    TechnicalItem("donusum_param", "Dönüşüm Parametresi", "Adet", 2482.0),
]

OTHER_TECH_ITEMS = [
    TechnicalItem("takeometrik", "Takeometrik Ölçüm Karneleri", "Sayfa", 143.0),
    TechnicalItem("aplikasyon", "Aplikasyon Krokisi", "Sayfa", 143.0),
    TechnicalItem("ebatli_roperli", "Ebatlı Kroki ve Röperli Kroki", "Sayfa", 143.0),
    TechnicalItem("mat_50x70", "Mat Kopya (50x70)", "Adet", 518.0),
    TechnicalItem("mat_70x100", "Mat Kopya (70x100)", "Adet", 720.0),
    TechnicalItem("seffaf_50x70", "Şeffaf Kopya (50x70)", "Adet", 720.0),
    TechnicalItem("seffaf_70x100", "Şeffaf Kopya (70x100)", "Adet", 955.0),
    TechnicalItem("tutga", "TUTGA Noktası", "Adet", 1773.0),
    TechnicalItem("c1_aga", "C1 Derece AGA Noktası", "Adet", 1773.0),
    TechnicalItem("c2_sga", "C2 Derece SGA Noktası", "Adet", 1033.0),
    TechnicalItem("c3_asn", "C3 Derece ASN Noktası", "Adet", 778.0),
    TechnicalItem("c4_nokta", "C4 Poligon/Fotogrametrik Nokta", "Adet", 338.0),
    TechnicalItem("ed50_ana", "ED-50 Ana Nirengi Noktası", "Adet", 665.0),
    TechnicalItem("ed50_ara", "ED-50 Ara Nirengi Noktası", "Adet", 338.0),
]


def corner_coordinate_fee(count: int) -> tuple[float, str]:
    if count <= 0:
        return 0.0, "0,00 TL / Nokta"
    first = min(count, 500)
    second = min(max(count - 500, 0), 500)
    third = max(count - 1000, 0)
    total = (first * 29.0) + (second * 24.0) + (third * 11.0)
    return total, "Kademeli: 1-500=29,00 | 501-1.000=24,00 | 1.001+=11,00 TL / Nokta"


class KadastroDesktopApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(APP_NAME)
        self.geometry("1320x840")
        self.minsize(1140, 740)

        self.settings = self.load_settings()
        self.selected_year = int(self.settings.get("selected_year", BASE_TARIFF_YEAR))
        if self.selected_year < 2020 or self.selected_year > BASE_TARIFF_YEAR:
            self.selected_year = BASE_TARIFF_YEAR
        self.control_rates = self.load_control_rates()
        self.rows = self.load_rows(self.selected_year)
        self.il_index = self.build_index(self.rows)
        self.display_name = resolve_display_name()
        saved_theme = self.settings.get("theme", "Aydınlık")
        if saved_theme == "Aydinlik":
            saved_theme = "Aydınlık"
        if saved_theme == "Karanlik":
            saved_theme = "Karanlık"
        self.theme_var = tk.StringVar(value=saved_theme)

        self.style = ttk.Style(self)
        self.style.theme_use("clam")
        self.theme_targets: list[tk.Widget] = []
        self.spinbox_targets: list[tk.Spinbox] = []
        self.current_palette: dict[str, str] = {}
        self.year_var = tk.StringVar(value=str(self.selected_year))

        self.build_layout()
        self.apply_app_icon()
        self.apply_theme()

    def apply_app_icon(self) -> None:
        icon_candidates = [
            DEFAULT_ICON_PATH,
            resource_path("icon.png"),
            Path("icon.png"),
        ]
        for path in icon_candidates:
            try:
                if not path or not path.exists():
                    continue
                self._app_icon_img = tk.PhotoImage(file=str(path))  # type: ignore[attr-defined]
                self.iconphoto(True, self._app_icon_img)  # type: ignore[attr-defined]
                return
            except Exception:
                continue

    def load_settings(self) -> dict:
        if SETTINGS_FILE.exists():
            try:
                return json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                return {}
        return {}

    def save_settings(self) -> None:
        SETTINGS_FILE.write_text(json.dumps(self.settings, ensure_ascii=False, indent=2), encoding="utf-8")

    def load_control_rates(self) -> dict[str, float]:
        merged = dict(CONTROL_RATE_DEFAULTS)
        raw = self.settings.get("control_rates", {})
        if isinstance(raw, dict):
            for key, default_value in CONTROL_RATE_DEFAULTS.items():
                try:
                    merged[key] = float(raw.get(key, default_value))
                except (TypeError, ValueError):
                    merged[key] = float(default_value)
        self.settings["control_rates"] = merged
        return merged

    def get_rate(self, key: str) -> float:
        return pick_rate(self.control_rates, key, CONTROL_RATE_DEFAULTS.get(key, 0.0))

    def load_rows(self, year: int | None = None) -> list[dict]:
        active_year = int(year if year is not None else self.selected_year)
        base_rows = json.loads(resource_path("katsayilar.json").read_text(encoding="utf-8"))
        legacy_custom_rows = self.settings.get("custom_rows", [])
        custom_by_year = self.settings.get("custom_rows_by_year", {})
        year_custom_rows = custom_by_year.get(str(active_year), [])

        merged: dict[tuple[str, str], dict] = {}
        for row in base_rows:
            il = normalize_key(str(row.get("il", "")))
            ilce = normalize_key(str(row.get("ilce", "")))
            if il and ilce:
                merged[(il, ilce)] = row

        for row in legacy_custom_rows + year_custom_rows:
            il = normalize_key(str(row.get("il", "")))
            ilce = normalize_key(str(row.get("ilce", "")))
            if il and ilce:
                merged[(il, ilce)] = row

        return list(merged.values())

    def rebuild_data(self, year: int | None = None) -> None:
        if year is not None:
            self.selected_year = int(year)
            self.year_var.set(str(self.selected_year))
        self.rows = self.load_rows(self.selected_year)
        self.il_index = self.build_index(self.rows)

    def get_year_options(self) -> list[int]:
        return list(range(BASE_TARIFF_YEAR, 2019, -1))

    @staticmethod
    def build_index(rows: list[dict]) -> dict[str, dict[str, dict[str, float | str]]]:
        index: dict[str, dict[str, dict[str, float | str]]] = {}
        for row in rows:
            il = normalize_key(str(row.get("il", "")))
            ilce = normalize_key(str(row.get("ilce", "")))
            if not il or not ilce:
                continue
            index.setdefault(il, {})[ilce] = {
                "bolge": normalize_key(str(row.get("bolge", ""))),
                "tapu": float(row.get("tapu", 1.0)),
                "kadastro": float(row.get("kadastro", 1.0)),
            }
        return index

    def build_layout(self) -> None:
        top = ttk.Frame(self, padding=(14, 12, 14, 10))
        top.pack(fill="x")

        title_block = ttk.Frame(top)
        title_block.pack(side="left", fill="x", expand=True)
        ttk.Label(
            title_block,
            text=f"Hoşgeldiniz {self.display_name}",
            style="Title.TLabel",
        ).pack(side="left", padx=(0, 14))

        theme_block = ttk.Frame(top)
        theme_block.pack(side="right")
        ttk.Label(theme_block, text="Aktif Katsayı Yılı", style="Muted.TLabel").pack(side="left", padx=(0, 6))
        self.year_box = ttk.Combobox(
            theme_block,
            textvariable=self.year_var,
            values=[str(y) for y in self.get_year_options()],
            width=8,
            state="readonly",
        )
        self.year_box.pack(side="left", padx=(0, 12))
        self.year_box.bind("<<ComboboxSelected>>", self.on_year_change)

        ttk.Label(theme_block, text="Tema", style="Muted.TLabel").pack(side="left", padx=(0, 6))
        self.theme_box = ttk.Combobox(
            theme_block,
            textvariable=self.theme_var,
            values=["Aydınlık", "Karanlık"],
            width=12,
            state="readonly",
        )
        self.theme_box.pack(side="left")
        self.theme_box.bind("<<ComboboxSelected>>", self.on_theme_change)

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        self.tab_technical = ttk.Frame(self.notebook, padding=14)
        self.tab_control = ttk.Frame(self.notebook, padding=14)
        self.tab_admin = ttk.Frame(self.notebook, padding=14)
        self.tab_about = ttk.Frame(self.notebook, padding=14)
        self.notebook.add(self.tab_technical, text="Teknik Bilgi Belge")
        self.notebook.add(self.tab_control, text="Kontrollük")
        self.notebook.add(self.tab_admin, text="⚙")
        self.notebook.add(self.tab_about, text="ℹ")

        self.build_technical_tab()
        self.build_control_tab()
        self.build_admin_tab()
        self.build_about_tab()

    def on_theme_change(self, _event: object | None = None) -> None:
        self.settings["theme"] = self.theme_var.get()
        self.save_settings()
        self.apply_theme()

    def on_year_change(self, _event: object | None = None) -> None:
        try:
            year = int(self.year_var.get())
        except ValueError:
            messagebox.showerror("Hata", "Geçersiz yıl seçimi.")
            return
        if year < 2020 or year > BASE_TARIFF_YEAR:
            year = BASE_TARIFF_YEAR
            self.year_var.set(str(year))
        self.settings["selected_year"] = year
        self.save_settings()
        self.rebuild_data(year)
        self.refresh_control_iller()
        self.refresh_admin_ui()
        if year != BASE_TARIFF_YEAR:
            messagebox.showwarning("Uyarı", f"{year} yılı için katsayılar henüz yüklenmedi.")

    def apply_theme(self) -> None:
        palettes = {
            "Aydınlık": {
                "bg": "#eef3fb",
                "panel": "#ffffff",
                "panel_soft": "#f8fbff",
                "fg": "#0f172a",
                "input_bg": "#ffffff",
                "input_fg": "#0b1220",
                "accent": "#2563eb",
                "accent_hover": "#1d4ed8",
                "muted": "#5f6f89",
                "tree_bg": "#ffffff",
                "border": "#c8d6eb",
                "danger": "#dc2626",
            },
            "Karanlık": {
                "bg": "#070d1a",
                "panel": "#111a2b",
                "panel_soft": "#16233a",
                "fg": "#edf3ff",
                "input_bg": "#1a2a43",
                "input_fg": "#f8fbff",
                "accent": "#3f7fb8",
                "accent_hover": "#366d9e",
                "muted": "#9db0d1",
                "tree_bg": "#152238",
                "border": "#2a3e62",
                "danger": "#f87171",
            },
        }
        selected = self.theme_var.get()
        if selected == "Aydinlik":
            selected = "Aydınlık"
        if selected == "Karanlik":
            selected = "Karanlık"
        p = palettes.get(selected, palettes["Aydınlık"])
        self.current_palette = p

        self.configure(bg=p["bg"])
        self.style.configure("TFrame", background=p["bg"])
        self.style.configure("TLabel", background=p["bg"], foreground=p["fg"])
        self.style.configure("TLabelframe", background=p["bg"], foreground=p["fg"])
        self.style.configure("TLabelframe.Label", background=p["bg"], foreground=p["fg"])
        self.style.configure("Title.TLabel", font=("Segoe UI", 15, "bold"), background=p["bg"], foreground=p["fg"])
        self.style.configure("Muted.TLabel", background=p["bg"], foreground=p["muted"], font=("Segoe UI", 9))
        self.style.configure("Info.TLabel", background=p["bg"], foreground=p["fg"], font=("Segoe UI", 10, "bold"))

        self.style.configure("TButton", background=p["panel"], foreground=p["fg"], padding=(10, 7))
        self.style.map("TButton", background=[("active", p["panel_soft"])], foreground=[("active", p["fg"])])
        self.style.configure("Primary.TButton", background=p["accent"], foreground="#ffffff", padding=(10, 8))
        self.style.map(
            "Primary.TButton",
            background=[("active", p["accent_hover"]), ("pressed", p["accent_hover"])],
            foreground=[("active", "#ffffff"), ("pressed", "#ffffff")],
        )
        self.style.configure("Danger.TButton", background=p["danger"], foreground="#ffffff", padding=(10, 8))
        self.style.map(
            "Danger.TButton",
            background=[("active", p["danger"]), ("pressed", p["danger"])],
            foreground=[("active", "#ffffff"), ("pressed", "#ffffff")],
        )
        self.style.configure("Operation.TButton", anchor="w", background=p["panel"], foreground=p["fg"], padding=(12, 9))
        self.style.configure("OperationActive.TButton", anchor="w", background=p["accent"], foreground="#ffffff", padding=(12, 9))

        self.style.configure("TNotebook", background=p["bg"])
        self.style.configure("TNotebook.Tab", background=p["panel"], foreground=p["fg"], padding=(12, 7))
        self.style.map("TNotebook.Tab", background=[("selected", p["accent"])], foreground=[("selected", "#ffffff")])
        self.style.configure("TEntry", fieldbackground=p["input_bg"], foreground=p["input_fg"])
        self.style.configure("TCombobox", fieldbackground=p["input_bg"], background=p["input_bg"], foreground=p["input_fg"])
        self.style.map(
            "TCombobox",
            fieldbackground=[("readonly", p["input_bg"])],
            foreground=[("readonly", p["input_fg"])],
            selectbackground=[("readonly", p["accent"])],
            selectforeground=[("readonly", "#ffffff")],
        )
        self.style.configure("TSpinbox", fieldbackground=p["input_bg"], background=p["input_bg"], foreground=p["input_fg"])
        self.style.configure("Treeview", background=p["tree_bg"], fieldbackground=p["tree_bg"], foreground=p["fg"])
        self.style.map("Treeview", background=[("selected", p["accent"])], foreground=[("selected", "#ffffff")])
        self.style.configure("Treeview.Heading", background=p["panel"], foreground=p["fg"], font=("Segoe UI", 9, "bold"))

        self.option_add("*TCombobox*Listbox.background", p["panel"])
        self.option_add("*TCombobox*Listbox.foreground", p["fg"])
        self.option_add("*TCombobox*Listbox.selectBackground", p["accent"])
        self.option_add("*TCombobox*Listbox.selectForeground", "#ffffff")

        for widget in self.theme_targets:
            try:
                if not widget.winfo_exists():
                    continue
                if isinstance(widget, tk.Text):
                    widget.configure(bg=p["input_bg"], fg=p["input_fg"], insertbackground=p["input_fg"])
                if isinstance(widget, tk.Frame):
                    widget.configure(bg=p["bg"])
            except tk.TclError:
                continue

        if hasattr(self, "about_card") and isinstance(self.about_card, tk.Frame):
            try:
                if self.about_card.winfo_exists():
                    self.about_card.configure(bg=p["panel"], highlightbackground=p["border"])
            except tk.TclError:
                pass

        for spin in self.spinbox_targets:
            spin.configure(
                bg=p["input_bg"],
                fg=p["input_fg"],
                insertbackground=p["input_fg"],
                buttonbackground=p["accent"],
                highlightbackground=p["border"],
                highlightcolor=p["accent"],
                disabledbackground=p["input_bg"],
                readonlybackground=p["input_bg"],
                relief="solid",
                borderwidth=2,
            )

    def build_technical_tab(self) -> None:
        wrapper = ttk.Frame(self.tab_technical)
        wrapper.pack(fill="both", expand=True)

        intro = ttk.Frame(wrapper)
        intro.pack(fill="x", pady=(0, 10))
        ttk.Label(intro, text="Teknik Bilgi-Belge Hesaplama", style="Info.TLabel").pack(side="left", anchor="w")
        self.mahkeme_infaz_label = ttk.Label(intro, text="", style="Muted.TLabel")
        self.mahkeme_infaz_label.pack(side="left", padx=(14, 0))
        self.refresh_mahkeme_infaz_label()

        content = ttk.Frame(wrapper)
        content.pack(fill="both", expand=True)
        content.grid_columnconfigure(0, weight=1)
        content.grid_rowconfigure(0, weight=1)

        quick = ttk.Frame(content, padding=12)
        quick.grid(row=0, column=0, sticky="nsew", padx=(0, 12))
        quick.grid_columnconfigure(1, weight=1)

        headers = ["Sıra", "Kalem", "Birim Fiyat", "Miktar", "Tutar"]
        for idx, text in enumerate(headers):
            ttk.Label(quick, text=text, font=("Segoe UI", 10, "bold")).grid(row=0, column=idx, sticky="w", padx=6, pady=(0, 10))

        self.tech_qty_vars: dict[str, tk.IntVar] = {
            "parcel_corner": tk.IntVar(value=0),
            "poligon_xy": tk.IntVar(value=0),
            "olcu_kroki": tk.IntVar(value=0),
            "donusum_param": tk.IntVar(value=0),
        }
        self.tech_line_total_labels: dict[str, ttk.Label] = {}
        self.corner_price_label = ttk.Label(quick, text="29,00 / 24,00 / 11,00 TL / Nokta")

        main_map = {item.key: item for item in MAIN_TECH_ITEMS}
        row = 1
        for order, key in enumerate(["parcel_corner", "poligon_xy", "olcu_kroki", "donusum_param"], start=1):
            item = main_map[key]
            ttk.Label(quick, text=str(order)).grid(row=row, column=0, sticky="w", padx=6, pady=4)
            ttk.Label(quick, text=item.label).grid(row=row, column=1, sticky="w", padx=6, pady=4)
            if key == "parcel_corner":
                self.corner_price_label.grid(row=row, column=2, sticky="w", padx=6, pady=4)
            else:
                ttk.Label(quick, text=f"{tr_money(item.price)} / {item.unit}").grid(row=row, column=2, sticky="w", padx=6, pady=4)
            spin = tk.Spinbox(
                quick,
                from_=0,
                to=500000,
                textvariable=self.tech_qty_vars[key],
                width=10,
                increment=1,
                justify="center",
                font=("Segoe UI", 12, "bold"),
                relief="solid",
                borderwidth=2,
            )
            spin.grid(row=row, column=3, sticky="w", padx=6, pady=4)
            self.spinbox_targets.append(spin)
            total_label = ttk.Label(quick, text=tr_money(0))
            total_label.grid(row=row, column=4, sticky="w", padx=6, pady=4)
            self.tech_line_total_labels[key] = total_label
            row += 1

        ttk.Label(quick, text="5").grid(row=row, column=0, sticky="w", padx=6, pady=(8, 4))
        ttk.Label(quick, text="Diğer Teknik Bilgi-Belgeler").grid(row=row, column=1, sticky="w", padx=6, pady=(8, 4))
        self.other_item_var = tk.StringVar(value="--")
        self.other_qty_var = tk.IntVar(value=1)
        other_names = ["--"] + [item.label for item in OTHER_TECH_ITEMS]
        self.other_item_box = ttk.Combobox(quick, textvariable=self.other_item_var, values=other_names, state="readonly", width=38)
        self.other_item_box.grid(row=row, column=2, sticky="w", padx=6, pady=(8, 4))
        self.other_qty_spin = tk.Spinbox(
            quick,
            from_=1,
            to=100000,
            textvariable=self.other_qty_var,
            width=10,
            increment=1,
            justify="center",
            font=("Segoe UI", 12, "bold"),
            relief="solid",
            borderwidth=2,
        )
        self.other_qty_spin.grid(row=row, column=3, sticky="w", padx=6, pady=(8, 4))
        self.spinbox_targets.append(self.other_qty_spin)
        ttk.Button(quick, text="Ekle", style="Primary.TButton", command=self.add_other_item).grid(
            row=row,
            column=4,
            sticky="w",
            padx=6,
            pady=(8, 4),
        )

        self.other_selected_entries: list[dict] = []
        other_frame = ttk.Frame(quick)
        other_frame.grid(row=row + 1, column=1, columnspan=4, sticky="nsew", padx=6, pady=(6, 4))
        other_cols = ("kalem", "birim", "miktar", "tutar")
        self.other_tree = ttk.Treeview(other_frame, columns=other_cols, show="headings", height=6)
        for col, title, width in [
            ("kalem", "Kalem", 360),
            ("birim", "Birim Fiyat", 140),
            ("miktar", "Miktar", 90),
            ("tutar", "Tutar", 140),
        ]:
            self.other_tree.heading(col, text=title)
            self.other_tree.column(col, width=width, anchor="w")
        self.other_tree.pack(side="left", fill="x", expand=True)
        other_scroll = ttk.Scrollbar(other_frame, orient="vertical", command=self.other_tree.yview)
        other_scroll.pack(side="left", fill="y")
        self.other_tree.configure(yscrollcommand=other_scroll.set)
        other_actions = ttk.Frame(quick)
        other_actions.grid(row=row + 2, column=1, columnspan=4, sticky="w", padx=6, pady=(2, 6))
        ttk.Button(other_actions, text="Seçili Diğeri Sil", command=self.remove_other_item).pack(side="left")

        self.tech_side_panel = ttk.LabelFrame(content, text="İşlem Paneli", padding=14)
        self.tech_side_panel.grid(row=0, column=1, sticky="ns")
        self.tech_calc_button = ttk.Button(
            self.tech_side_panel,
            text="Teknik Bilgi-Belge Hesapla",
            style="Primary.TButton",
            command=self.calculate_technical,
        )
        self.tech_calc_button.pack(fill="x", pady=(0, 12))
        self.tech_total_caption = ttk.Label(self.tech_side_panel, text="Toplam Tutar: -", font=("Segoe UI", 11, "bold"))
        self.tech_total_caption.pack(anchor="w", pady=(0, 16))
        self.btn_pdf_save = ttk.Button(
            self.tech_side_panel,
            text="Çıktı Al (PDF)",
            command=self.save_technical_pdf,
            state="disabled",
        )
        self.btn_pdf_save.pack(fill="x", pady=(0, 8))
        self.btn_clear_technical = ttk.Button(
            self.tech_side_panel,
            text="Temizle",
            style="Danger.TButton",
            command=self.reset_technical,
            state="disabled",
        )
        self.btn_clear_technical.pack(fill="x")
        ttk.Label(
            self.tech_side_panel,
            text="PDF çıktısı örnek teknik bilgi-belge formatına yakın üretilir.",
            style="Muted.TLabel",
            wraplength=220,
            justify="left",
        ).pack(anchor="w", pady=(14, 0))

        details_box = ttk.LabelFrame(wrapper, text="Hesaplama Sonucu", padding=8)
        details_box.pack(fill="both", expand=True, pady=(10, 0))
        self.technical_details = tk.Text(details_box, height=13, wrap="word", relief="flat", borderwidth=0, padx=8, pady=8)
        self.technical_details.pack(fill="both", expand=True)
        self.theme_targets.append(self.technical_details)
        self.technical_details.insert("1.0", "Hesaplama sonucu burada görünecek.")
        self.last_technical_report: dict | None = None
        self.last_pdf_form_data: dict[str, object] = {}

    def refresh_mahkeme_infaz_label(self) -> None:
        if hasattr(self, "mahkeme_infaz_label"):
            value = self.get_rate("mahkeme_infazi")
            self.mahkeme_infaz_label.configure(text=f"Mahkeme Kararları İnfazı: {tr_money(value)}")

    def add_other_item(self) -> None:
        label = self.other_item_var.get().strip()
        qty = max(int(self.other_qty_var.get() or 0), 0)
        if not label or label == "--" or qty <= 0:
            messagebox.showerror("Hata", "Diğer kalem ve miktar seçiniz.")
            return
        item_map = {item.label: item for item in OTHER_TECH_ITEMS}
        item = item_map.get(label)
        if not item:
            messagebox.showerror("Hata", "Geçersiz diğer kalem.")
            return
        self.other_selected_entries.append({"item": item, "qty": qty})
        self.refresh_other_tree()

    def remove_other_item(self) -> None:
        selected = self.other_tree.selection()
        if not selected:
            return
        idx = self.other_tree.index(selected[0])
        if 0 <= idx < len(self.other_selected_entries):
            self.other_selected_entries.pop(idx)
            self.refresh_other_tree()

    def clear_other_items(self) -> None:
        self.other_selected_entries.clear()
        self.refresh_other_tree()

    def refresh_other_tree(self) -> None:
        for row_id in self.other_tree.get_children():
            self.other_tree.delete(row_id)
        for entry in self.other_selected_entries:
            item = entry["item"]
            qty = int(entry["qty"])
            total = qty * item.price
            self.other_tree.insert("", "end", values=(item.label, tr_money(item.price), qty, tr_money(total)))

    def reset_technical(self) -> None:
        for var in self.tech_qty_vars.values():
            var.set(0)
        self.other_qty_var.set(1)
        self.other_item_var.set("--")
        self.other_selected_entries.clear()
        self.refresh_other_tree()
        self.tech_total_caption.configure(text="Toplam Tutar: -")
        self.technical_details.delete("1.0", "end")
        self.technical_details.insert("1.0", "Hesaplama sonucu burada görünecek.")
        for label in self.tech_line_total_labels.values():
            label.configure(text=tr_money(0))
        self.corner_price_label.configure(text="29,00 / 24,00 / 11,00 TL / Nokta")
        self.btn_pdf_save.configure(state="disabled")
        self.btn_clear_technical.configure(state="disabled")
        self.last_technical_report = None

    def collect_technical_rows(self) -> tuple[list[dict], float]:
        rows: list[dict] = []
        total = 0.0

        corner_qty = max(int(self.tech_qty_vars["parcel_corner"].get() or 0), 0)
        corner_total, corner_price_label = corner_coordinate_fee(corner_qty)
        self.corner_price_label.configure(text=corner_price_label)
        rows.append(
            {
                "sira": 1,
                "cinsi": "Parsel Köşe Noktası",
                "adet": corner_qty,
                "birim": corner_price_label,
                "tutar": corner_total,
            }
        )
        self.tech_line_total_labels["parcel_corner"].configure(text=tr_money(corner_total))
        total += corner_total

        for order, key in [(2, "poligon_xy"), (3, "olcu_kroki"), (4, "donusum_param")]:
            item = next(i for i in MAIN_TECH_ITEMS if i.key == key)
            qty = max(int(self.tech_qty_vars[key].get() or 0), 0)
            line_total = qty * item.price
            rows.append(
                {
                    "sira": order,
                    "cinsi": item.label,
                    "adet": qty,
                    "birim": f"{tr_money(item.price)} / {item.unit}",
                    "tutar": line_total,
                }
            )
            self.tech_line_total_labels[key].configure(text=tr_money(line_total))
            total += line_total

        other_order = 5
        for entry in self.other_selected_entries:
            item = entry["item"]
            qty = max(int(entry["qty"]), 0)
            line_total = qty * item.price
            rows.append(
                {
                    "sira": other_order,
                    "cinsi": f"Diğer - {item.label}",
                    "adet": qty,
                    "birim": f"{tr_money(item.price)} / {item.unit}",
                    "tutar": line_total,
                }
            )
            total += line_total
            other_order += 1

        return rows, total

    def calculate_technical(self) -> None:
        rows, total = self.collect_technical_rows()
        rounded_total = round_lira(total)
        self.tech_total_caption.configure(text=f"Toplam Tutar: {tr_money(rounded_total)}")

        selected_il = self.control_il_var.get() if hasattr(self, "control_il_var") else ""
        if not selected_il:
            selected_il = normalize_key(self.settings.get("default_il", ""))
        if not selected_il and self.il_index:
            selected_il = sorted(self.il_index.keys())[0]
        selected_ilce = self.control_ilce_var.get() if hasattr(self, "control_ilce_var") else ""
        mudurluk = mudurluk_name_from_il(selected_il)

        lines = [
            "Teknik Bilgi-Belge Hesaplama Sonucu",
            f"Tarih: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            f"Müdürlük: {mudurluk}",
            f"Bölge: {selected_il} / {selected_ilce}".strip(" /"),
            f"Toplam Bedel: {tr_money(rounded_total)}",
            "",
            "Kalemler:",
        ]
        for row in rows:
            lines.append(
                f"- {row['sira']}. {row['cinsi']} | {row['adet']} | {row['birim']} | {tr_money(row['tutar'])}"
            )
        self.technical_details.delete("1.0", "end")
        self.technical_details.insert("1.0", "\n".join(lines))

        self.last_technical_report = {
            "date": datetime.now(),
            "rows": rows,
            "total": total,
            "il": selected_il,
            "ilce": selected_ilce,
            "mudurluk": mudurluk,
        }
        self.btn_pdf_save.configure(state="normal")
        self.btn_clear_technical.configure(state="normal")

    def resolve_pdf_fonts(self) -> tuple[str, str]:
        if hasattr(self, "_cached_pdf_fonts"):
            return self._cached_pdf_fonts  # type: ignore[attr-defined]
        try:
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except Exception:
            self._cached_pdf_fonts = ("Helvetica", "Helvetica-Bold")
            return self._cached_pdf_fonts  # type: ignore[attr-defined]

        font_candidates = [
            (
                resource_path("fonts/DejaVuSans.ttf"),
                resource_path("fonts/DejaVuSans-Bold.ttf"),
            ),
            (
                Path("C:/Windows/Fonts/arial.ttf"),
                Path("C:/Windows/Fonts/arialbd.ttf"),
            ),
            (
                Path("C:/Windows/Fonts/calibri.ttf"),
                Path("C:/Windows/Fonts/calibrib.ttf"),
            ),
        ]
        for normal_file, bold_file in font_candidates:
            if not normal_file.exists() or not bold_file.exists():
                continue
            try:
                normal_name = "KadastroPDF-Regular"
                bold_name = "KadastroPDF-Bold"
                registered = set(pdfmetrics.getRegisteredFontNames())
                if normal_name not in registered:
                    pdfmetrics.registerFont(TTFont(normal_name, str(normal_file)))
                if bold_name not in registered:
                    pdfmetrics.registerFont(TTFont(bold_name, str(bold_file)))
                self._cached_pdf_fonts = (normal_name, bold_name)
                return self._cached_pdf_fonts  # type: ignore[attr-defined]
            except Exception:
                continue

        self._cached_pdf_fonts = ("Helvetica", "Helvetica-Bold")
        return self._cached_pdf_fonts  # type: ignore[attr-defined]

    def resolve_logo_path(self) -> Path | None:
        candidates = [
            Path(self.settings.get("logo_path", "")).expanduser() if self.settings.get("logo_path") else None,
            DEFAULT_LOGO_PATH,
            resource_path("assets/Tapu ve Kadastro Yeni Logo.jpg"),
        ]
        for candidate in candidates:
            if candidate and candidate.exists():
                return candidate
        return None

    def default_pdf_form_data(self) -> dict[str, object]:
        report = self.last_technical_report or {}
        stored_rows = self.last_pdf_form_data.get("parcel_rows", [])
        parcel_rows = self.normalize_pdf_parcel_rows(stored_rows)
        if not parcel_rows:
            parcel_rows = [
                {
                    "mahalle_koy": self.last_pdf_form_data.get("mahalle_koy", ""),
                    "pafta_no": self.last_pdf_form_data.get("pafta_no", ""),
                    "ada_no": self.last_pdf_form_data.get("ada_no", ""),
                    "parsel_no": self.last_pdf_form_data.get("parsel_no", ""),
                    "yuzolcumu": self.last_pdf_form_data.get("yuzolcumu", ""),
                }
            ]
        return {
            "parcel_rows": parcel_rows,
            "arz_oncesi": self.last_pdf_form_data.get("arz_oncesi", ""),
            "harita_kadastro_muh": self.last_pdf_form_data.get("harita_kadastro_muh", ""),
            "tasinmaz_sahibi": self.last_pdf_form_data.get("tasinmaz_sahibi", ""),
            "teslim_eden": self.last_pdf_form_data.get("teslim_eden", ""),
            "teslim_alan": self.last_pdf_form_data.get("teslim_alan", ""),
            "mudurluk": str(report.get("mudurluk") or ""),
        }

    @staticmethod
    def normalize_pdf_parcel_rows(rows: object) -> list[dict[str, str]]:
        out: list[dict[str, str]] = []
        if not isinstance(rows, list):
            return out
        for row in rows:
            if not isinstance(row, dict):
                continue
            normalized = {
                "mahalle_koy": str(row.get("mahalle_koy", "")).strip(),
                "pafta_no": str(row.get("pafta_no", "")).strip(),
                "ada_no": str(row.get("ada_no", "")).strip(),
                "parsel_no": str(row.get("parsel_no", "")).strip(),
                "yuzolcumu": str(row.get("yuzolcumu", "")).strip(),
            }
            if any(normalized.values()):
                out.append(normalized)
        return out

    def open_pdf_form_dialog(self) -> None:
        if not self.last_technical_report:
            messagebox.showerror("Hata", "Önce hesaplama yapınız.")
            return

        old_window = getattr(self, "pdf_form_window", None)
        if old_window and old_window.winfo_exists():
            old_window.lift()
            old_window.focus_force()
            return

        dialog = tk.Toplevel(self)
        self.pdf_form_window = dialog
        dialog.title("PDF Önizleme ve Form Alanları")
        dialog.geometry("920x640")
        dialog.minsize(860, 590)
        dialog.transient(self)
        dialog.grab_set()

        p = self.current_palette or {"bg": "#0b1020", "panel": "#111a30", "fg": "#f1f5f9", "border": "#2a3e62", "muted": "#9db0d1"}
        dialog.configure(bg=p["bg"])

        outer = ttk.Frame(dialog, padding=12)
        outer.pack(fill="both", expand=True)
        ttk.Label(outer, text="PDF Çıktı Önizleme", style="Info.TLabel").pack(anchor="w")
        ttk.Label(
            outer,
            text="Bu alandaki bilgiler opsiyoneldir. Girilmese de sistem PDF üretir.",
            style="Muted.TLabel",
        ).pack(anchor="w", pady=(2, 10))

        card = tk.Frame(outer, bg=p["panel"], highlightbackground=p["border"], highlightthickness=1, bd=0)
        card.pack(fill="both", expand=True)
        self.theme_targets.append(card)

        form_wrap = ttk.Frame(card, padding=12)
        form_wrap.pack(fill="both", expand=True)

        defaults = self.default_pdf_form_data()
        vars_data = {
            key: tk.StringVar(value=str(value))
            for key, value in defaults.items()
            if key not in {"parcel_rows"}
        }
        self._pdf_form_vars = vars_data
        self._pdf_row_vars: list[dict[str, tk.StringVar]] = []

        top_section = ttk.LabelFrame(form_wrap, text="Taşınmaz Bilgileri", padding=10)
        top_section.pack(fill="x")
        row_fields = [("MAHALLE / KÖY", 30), ("PAFTA NO", 12), ("ADA NO", 12), ("PARSEL NO", 12), ("YÜZÖLÇÜMÜ", 14)]
        for col, (label, _) in enumerate(row_fields):
            ttk.Label(top_section, text=label).grid(row=0, column=col, sticky="w", padx=(0, 8), pady=(0, 4))
        row_host = ttk.Frame(top_section)
        row_host.grid(row=1, column=0, columnspan=5, sticky="w")
        self._pdf_rows_host = row_host

        def add_row(default_row: dict[str, str] | None = None) -> None:
            host = self._pdf_rows_host
            row_index = len(self._pdf_row_vars)
            defaults_row = default_row or {}
            row_vars = {
                "mahalle_koy": tk.StringVar(value=str(defaults_row.get("mahalle_koy", ""))),
                "pafta_no": tk.StringVar(value=str(defaults_row.get("pafta_no", ""))),
                "ada_no": tk.StringVar(value=str(defaults_row.get("ada_no", ""))),
                "parsel_no": tk.StringVar(value=str(defaults_row.get("parsel_no", ""))),
                "yuzolcumu": tk.StringVar(value=str(defaults_row.get("yuzolcumu", ""))),
            }
            ttk.Entry(host, textvariable=row_vars["mahalle_koy"], width=30).grid(row=row_index, column=0, sticky="w", padx=(0, 8), pady=3)
            ttk.Entry(host, textvariable=row_vars["pafta_no"], width=12).grid(row=row_index, column=1, sticky="w", padx=(0, 8), pady=3)
            ttk.Entry(host, textvariable=row_vars["ada_no"], width=12).grid(row=row_index, column=2, sticky="w", padx=(0, 8), pady=3)
            ttk.Entry(host, textvariable=row_vars["parsel_no"], width=12).grid(row=row_index, column=3, sticky="w", padx=(0, 8), pady=3)
            ttk.Entry(host, textvariable=row_vars["yuzolcumu"], width=14).grid(row=row_index, column=4, sticky="w", padx=(0, 8), pady=3)
            self._pdf_row_vars.append(row_vars)

        for row in defaults.get("parcel_rows", []) if isinstance(defaults.get("parcel_rows"), list) else []:
            add_row(row if isinstance(row, dict) else None)
        if not self._pdf_row_vars:
            add_row()

        row_actions = ttk.Frame(top_section)
        row_actions.grid(row=2, column=0, columnspan=5, sticky="w", pady=(6, 0))
        ttk.Button(row_actions, text="+", width=4, style="Primary.TButton", command=lambda: add_row()).pack(side="left")

        arz_section = ttk.LabelFrame(form_wrap, text="Talep Metni", padding=10)
        arz_section.pack(fill="x", pady=(10, 0))
        ttk.Label(arz_section, text="Metin:").grid(row=0, column=0, sticky="w")
        ttk.Entry(arz_section, textvariable=vars_data["arz_oncesi"], width=86).grid(row=1, column=0, sticky="w", pady=(4, 0))

        sign_section = ttk.LabelFrame(form_wrap, text="İmza ve Sorumlu Alanları", padding=10)
        sign_section.pack(fill="x", pady=(10, 0))

        sign_fields = [
            ("Sorumlu Harita Kadastro Mühendisi", "harita_kadastro_muh"),
            ("Taşınmaz Sahibi / Vekili", "tasinmaz_sahibi"),
            ("TESLİM EDEN", "teslim_eden"),
            ("TESLİM ALAN", "teslim_alan"),
        ]
        for idx, (label, key) in enumerate(sign_fields):
            ttk.Label(sign_section, text=label).grid(row=idx, column=0, sticky="w", padx=(0, 8), pady=4)
            ttk.Entry(sign_section, textvariable=vars_data[key], width=46).grid(row=idx, column=1, sticky="w", padx=(0, 8), pady=4)

        footer = ttk.Frame(form_wrap)
        footer.pack(fill="x", pady=(14, 0))
        ttk.Button(footer, text="Yazdır ve PDF Kaydet", style="Primary.TButton", command=self.save_pdf_from_dialog).pack(side="left")
        ttk.Button(footer, text="İptal", command=dialog.destroy).pack(side="left", padx=8)

    def save_pdf_from_dialog(self) -> None:
        if not self.last_technical_report:
            messagebox.showerror("Hata", "Önce hesaplama yapınız.")
            return
        vars_data = getattr(self, "_pdf_form_vars", {})
        form_data = {k: str(v.get()).strip() for k, v in vars_data.items() if hasattr(v, "get")}
        parcel_rows: list[dict[str, str]] = []
        for row_vars in getattr(self, "_pdf_row_vars", []):
            row = {
                "mahalle_koy": row_vars["mahalle_koy"].get().strip(),
                "pafta_no": row_vars["pafta_no"].get().strip(),
                "ada_no": row_vars["ada_no"].get().strip(),
                "parsel_no": row_vars["parsel_no"].get().strip(),
                "yuzolcumu": row_vars["yuzolcumu"].get().strip(),
            }
            if any(row.values()):
                parcel_rows.append(row)
        form_data["parcel_rows"] = parcel_rows
        self.last_pdf_form_data = form_data

        default_name = f"teknik_bilgi_belge_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        selected = filedialog.asksaveasfilename(
            title="PDF Çıktı Kaydet",
            defaultextension=".pdf",
            initialfile=default_name,
            filetypes=[("PDF", "*.pdf")],
        )
        if not selected:
            return
        try:
            self.create_technical_pdf(Path(selected), form_data=form_data)
            messagebox.showinfo("Başarılı", f"PDF oluşturuldu:\n{selected}")
            win = getattr(self, "pdf_form_window", None)
            if win and win.winfo_exists():
                win.destroy()
        except Exception as exc:
            messagebox.showerror("Hata", str(exc))

    def create_technical_pdf(self, path: Path, form_data: dict[str, object] | None = None) -> None:
        if not self.last_technical_report:
            raise RuntimeError("Rapor oluşturmak için önce hesaplama yapınız.")
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfgen import canvas
            from reportlab.lib.utils import ImageReader
        except Exception as exc:
            raise RuntimeError(
                "PDF kütüphanesi eksik (reportlab). Lütfen `pip install reportlab` komutunu çalıştırın."
            ) from exc

        report = self.last_technical_report
        data = self.default_pdf_form_data()
        if form_data:
            data.update({k: str(v or "") for k, v in form_data.items()})
            if "parcel_rows" in form_data and isinstance(form_data["parcel_rows"], list):
                data["parcel_rows"] = form_data["parcel_rows"]  # type: ignore[index]

        parcel_rows = self.normalize_pdf_parcel_rows(data.get("parcel_rows", []))
        if not parcel_rows:
            parcel_rows = [{"mahalle_koy": "", "pafta_no": "", "ada_no": "", "parsel_no": "", "yuzolcumu": ""}]

        normal_font, bold_font = self.resolve_pdf_fonts()
        c = canvas.Canvas(str(path), pagesize=A4)
        width, height = A4
        left = 26
        right = width - 26
        top = height - 26
        bottom = 26

        c.setLineWidth(0.7)
        header_top = top - 2
        header_height = 94
        c.rect(left, header_top - header_height, right - left, header_height)
        logo_box_w = 150
        logo_left = left + 8
        logo_bottom = header_top - header_height + 8
        c.rect(logo_left, logo_bottom, logo_box_w, header_height - 16)
        logo_path = self.resolve_logo_path()
        if logo_path:
            try:
                img = ImageReader(str(logo_path))
                c.drawImage(
                    img,
                    logo_left + 4,
                    logo_bottom + 4,
                    width=logo_box_w - 8,
                    height=header_height - 24,
                    preserveAspectRatio=True,
                    anchor="c",
                    mask="auto",
                )
            except Exception:
                c.setFont(normal_font, 8.2)
                c.drawCentredString(logo_left + logo_box_w / 2, logo_bottom + (header_height - 16) / 2, "Logo")
        else:
            c.setFont(normal_font, 8.2)
            c.drawCentredString(logo_left + logo_box_w / 2, logo_bottom + (header_height - 16) / 2, "Logo")

        c.setFont(bold_font, 12.4)
        c.drawCentredString((left + right) / 2 + 72, header_top - 40, "KADASTRO BİLGİ VE BELGELERİ")
        c.drawCentredString((left + right) / 2 + 72, header_top - 58, "İSTEM BELGESİ VE TESLİM SENEDİ")

        info_top = header_top - header_height - 10
        parcel_row_count = max(4, len(parcel_rows))
        row_h = 18
        header_h = 22
        parcel_table_bottom = info_top - header_h - (parcel_row_count * row_h)
        info_height = header_h + (parcel_row_count * row_h) + 36 + 34
        info_bottom = info_top - info_height
        c.rect(left, info_bottom, right - left, info_height)
        c.line(left, info_top - header_h, right, info_top - header_h)
        for idx in range(parcel_row_count):
            y_line = info_top - header_h - ((idx + 1) * row_h)
            c.line(left, y_line, right, y_line)

        col_edges = [left, left + 170, left + 255, left + 336, left + 430, right]
        for x in col_edges[1:-1]:
            c.line(x, info_top, x, parcel_table_bottom)

        c.setFont(bold_font, 8.4)
        headers = ["MAHALLE / KÖY", "PAFTA NO", "ADA NO", "PARSEL NO", "YÜZÖLÇÜMÜ"]
        x_points = [left + 4, col_edges[1] + 4, col_edges[2] + 4, col_edges[3] + 4, col_edges[4] + 4]
        for head, x in zip(headers, x_points):
            c.drawString(x, info_top - 16, head)

        c.setFont(normal_font, 8.2)
        for idx, row in enumerate(parcel_rows[:parcel_row_count]):
            y_row = info_top - header_h - 14 - (idx * row_h)
            c.drawString(left + 4, y_row, row.get("mahalle_koy", ""))
            c.drawString(col_edges[1] + 4, y_row, row.get("pafta_no", ""))
            c.drawString(col_edges[2] + 4, y_row, row.get("ada_no", ""))
            c.drawString(col_edges[3] + 4, y_row, row.get("parsel_no", ""))
            c.drawString(col_edges[4] + 4, y_row, row.get("yuzolcumu", ""))

        request_prefix = str(data.get("arz_oncesi", "")).strip()
        request_line = "işleminde kullanacağımdan teknik bilgi ve belgenin tarafıma verilmesini arz ederim."
        if request_prefix:
            request_line = f"{request_prefix} {request_line}"
        request_line = request_line[:160]
        request_y = parcel_table_bottom - 16
        c.drawString(left + 4, request_y, request_line)
        c.drawRightString(right - 6, request_y, report["date"].strftime("%d/%m/%Y"))
        c.drawCentredString(left + 130, request_y - 14, "Sorumlu")
        c.drawCentredString(left + 130, request_y - 27, data.get("harita_kadastro_muh", "") or "Harita Kadastro Mühendisi")
        c.drawCentredString(right - 125, request_y - 14, "Taşınmaz Sahibi")
        c.drawCentredString(right - 125, request_y - 27, data.get("tasinmaz_sahibi", "") or "veya Vekili")

        table_top = info_bottom - 18
        c.setFont(bold_font, 10.2)
        c.drawString(left, table_top, "TESLİM SENEDİ")
        table_y = table_top - 8
        table_height = 318
        c.rect(left, table_y - table_height, right - left, table_height)

        column_edges = [left, left + 225, left + 280, left + 412, right]
        for edge in column_edges[1:-1]:
            c.line(edge, table_y, edge, table_y - table_height)
        c.line(left, table_y - 20, right, table_y - 20)

        c.setFont(bold_font, 8.8)
        col_headers = ["CİNSİ", "ADEDİ", "BİRİM FİYATI", "TOPLAM FİYATI"]
        for i, title in enumerate(col_headers):
            c.drawString(column_edges[i] + 4, table_y - 14, title)

        y = table_y - 35
        c.setFont(normal_font, 8.8)
        for row in report["rows"]:
            if y < table_y - table_height + 55:
                break
            c.drawString(column_edges[0] + 4, y, str(row["cinsi"])[:50])
            c.drawRightString(column_edges[2] - 6, y, str(row["adet"]))
            c.drawRightString(column_edges[3] - 6, y, str(row["birim"]))
            c.drawRightString(column_edges[4] - 6, y, tr_money(row["tutar"]))
            c.line(left, y - 4, right, y - 4)
            y -= 16

        total_line_y = table_y - table_height + 26
        c.setFont(bold_font, 10)
        c.drawString(left + 6, total_line_y, "Toplam")
        c.drawRightString(right - 8, total_line_y, tr_money(round_lira(report["total"])))

        note_y = table_y - table_height + 16
        c.setFont(normal_font, 8.1)
        c.drawString(
            left + 4,
            note_y,
            "Teknik belge bedeli tahsil edilerek tarafımızdan teslim ve tesellüm edilmiştir.",
        )

        sign_y = table_y - table_height - 24
        c.setFont(normal_font, 9)
        c.drawString(left + 12, sign_y, "TESLİM EDEN")
        c.drawString(left + 230, sign_y, "TESLİM ALAN")
        c.drawString(left + 430, sign_y, "GÖRÜLMÜŞTÜR")
        c.setFont(normal_font, 8.4)
        if data.get("teslim_eden"):
            c.drawString(left + 12, sign_y - 14, data["teslim_eden"])
        if data.get("teslim_alan"):
            c.drawString(left + 230, sign_y - 14, data["teslim_alan"])
        c.setFont(bold_font, 9)
        c.drawString(left + 420, sign_y - 14, data.get("mudurluk", str(report.get("mudurluk") or "Kadastro Müdürlüğü")))
        # Kullanıcı talebi doğrultusunda revizyon satırı PDF'te gizli tutulur.
        c.save()

    def save_technical_pdf(self) -> None:
        if not self.last_technical_report:
            messagebox.showerror("Hata", "Önce hesaplama yapınız.")
            return
        self.open_pdf_form_dialog()

    def build_control_tab(self) -> None:
        top = ttk.LabelFrame(self.tab_control, text="Bölge Seçimi", padding=10)
        top.pack(fill="x", pady=(0, 10))
        top.grid_columnconfigure(7, weight=1)

        ttk.Label(top, text="İl:").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.control_il_var = tk.StringVar()
        self.control_il_box = ttk.Combobox(top, textvariable=self.control_il_var, state="readonly", width=18)
        self.control_il_box.grid(row=0, column=1, sticky="w", padx=4, pady=4)
        self.control_il_box.bind("<<ComboboxSelected>>", self.on_control_il_change)

        ttk.Label(top, text="İlçe:").grid(row=0, column=2, sticky="w", padx=4, pady=4)
        self.control_ilce_var = tk.StringVar()
        self.control_ilce_box = ttk.Combobox(top, textvariable=self.control_ilce_var, state="readonly", width=24)
        self.control_ilce_box.grid(row=0, column=3, sticky="w", padx=4, pady=4)
        self.control_ilce_box.bind("<<ComboboxSelected>>", lambda _e: self.update_control_yk())

        self.yk_label = ttk.Label(top, text="Kadastro YK: -", font=("Segoe UI", 11, "bold"))
        self.yk_label.grid(row=0, column=4, sticky="w", padx=12, pady=4)
        self.mudurluk_preview_label = ttk.Label(top, text="Müdürlük: -", style="Muted.TLabel")
        self.mudurluk_preview_label.grid(row=0, column=5, sticky="w", padx=(8, 4), pady=4)

        body = ttk.Frame(self.tab_control)
        body.pack(fill="both", expand=True)
        left = ttk.LabelFrame(body, text="İşlem Türü", padding=10)
        left.pack(side="left", fill="y", padx=(0, 8))
        right = ttk.Frame(body)
        right.pack(side="left", fill="both", expand=True)

        self.control_type_var = tk.StringVar(value="degisiklik")
        self.operation_code = "degisiklik"
        self.operation_defs = [
            ("1- Değişiklik İşlemleri (15-16. Madde)", "degisiklik"),
            ("2- Kontrollük Hizmetleri (Parselasyon Planları)", "parselasyon"),
            ("3- Kamulaştırma Haritaları Kontrolü", "kamulastirma"),
            ("4- Mülkiyet Raporları", "mulkiyet"),
            ("5- İrtifak Hakkı Tesisi ve Terkini", "irtifak"),
            ("6- Birleştirme (Tevhit)", "birlestirme"),
            ("7- Cins Değişikliği", "cins"),
            ("8- Plan / Kroki Örneği", "plan_kroki"),
            ("9- Aplikasyon", "aplikasyon"),
            ("10- Parselin Yerinde Gösterilmesi", "parsel_yerinde"),
            ("11- Hatalı Bağımsız Bölüm/Blok Düzeltme", "hatali_bagimsiz"),
            ("12- ST/STK Harita ve Pafta Kopyaları", "st_stk"),
            ("13- Ülke Nirengi Koordinat Değerleri", "nirengi"),
        ]
        left_wrap = ttk.Frame(left)
        left_wrap.pack(fill="both", expand=True)
        self.operation_canvas = tk.Canvas(left_wrap, highlightthickness=0, borderwidth=0)
        op_scroll = ttk.Scrollbar(left_wrap, orient="vertical", command=self.operation_canvas.yview)
        self.operation_canvas.configure(yscrollcommand=op_scroll.set)
        self.operation_canvas.pack(side="left", fill="both", expand=True)
        op_scroll.pack(side="left", fill="y")
        operation_inner = ttk.Frame(self.operation_canvas)
        op_window = self.operation_canvas.create_window((0, 0), window=operation_inner, anchor="nw")

        def _sync_operation_scroll(_event: object | None = None) -> None:
            self.operation_canvas.configure(scrollregion=self.operation_canvas.bbox("all"))

        def _fit_operation_width(event: tk.Event) -> None:
            self.operation_canvas.itemconfigure(op_window, width=event.width)

        operation_inner.bind("<Configure>", _sync_operation_scroll)
        self.operation_canvas.bind("<Configure>", _fit_operation_width)

        def _on_op_wheel(event: tk.Event) -> None:
            if event.delta:
                self.operation_canvas.yview_scroll(int(-event.delta / 120), "units")
            elif getattr(event, "num", None) == 4:
                self.operation_canvas.yview_scroll(-1, "units")
            elif getattr(event, "num", None) == 5:
                self.operation_canvas.yview_scroll(1, "units")

        def _bind_wheel(_event: object | None = None) -> None:
            self.operation_canvas.bind_all("<MouseWheel>", _on_op_wheel)
            self.operation_canvas.bind_all("<Button-4>", _on_op_wheel)
            self.operation_canvas.bind_all("<Button-5>", _on_op_wheel)

        def _unbind_wheel(_event: object | None = None) -> None:
            self.operation_canvas.unbind_all("<MouseWheel>")
            self.operation_canvas.unbind_all("<Button-4>")
            self.operation_canvas.unbind_all("<Button-5>")

        self.operation_canvas.bind("<Enter>", _bind_wheel)
        self.operation_canvas.bind("<Leave>", _unbind_wheel)

        self.operation_buttons: dict[str, ttk.Button] = {}
        for title, code in self.operation_defs:
            btn = ttk.Button(
                operation_inner,
                text=title,
                style="Operation.TButton",
                command=lambda c=code: self.set_active_operation(c),
            )
            btn.pack(fill="x", pady=5)
            self.operation_buttons[code] = btn

        self.selected_process_label = ttk.Label(
            right,
            text="1- Değişiklik İşlemleri (15-16. Madde)",
            font=("Segoe UI", 11, "bold"),
        )
        self.selected_process_label.pack(anchor="w", pady=(0, 6))

        form = ttk.LabelFrame(right, text="Hesaplama için bilgileri aşağıya giriniz", padding=10)
        form.pack(fill="x", pady=(0, 8))

        self.area_var = tk.StringVar(value="")
        self.km_var = tk.StringVar(value="1")
        self.parsel_count_var = tk.IntVar(value=2)
        self.parsel_turu_var = tk.StringVar(value="Oluşmuş imar parselleri")
        self.plan_parsel_turu_var = tk.StringVar(value="İmar Parseli")
        self.cins_parsel_turu_var = tk.StringVar(value="İmar Parseli")
        self.cins_type_var = tk.StringVar(value="Yapısız iken yapılı hâle getirme (2.2.1)")
        self.cins_tarimsal_var = tk.BooleanVar(value=False)
        self.cins_fazla_yapi_var = tk.IntVar(value=0)
        self.cins_parsel_sayisi_var = tk.IntVar(value=1)
        self.cins_yapi_sayisi_var = tk.IntVar(value=1)
        self.cins_arsa_pay_orani_var = tk.StringVar(value="100")
        self.geriye_var = tk.BooleanVar(value=False)
        self.yeni_var = tk.BooleanVar(value=False)
        self.mera_var = tk.BooleanVar(value=False)
        self.include_docs_var = tk.BooleanVar(value=True)
        self.kam_subtype_var = tk.StringVar(value="Şeritvari bazlı")
        self.report_type_var = tk.StringVar(value="İmar ayırma/birleştirme (parsel)")
        self.report_parsel_var = tk.IntVar(value=1)
        self.hatali_bagimsiz_sayi_var = tk.IntVar(value=2)
        self.hatali_blok_sayi_var = tk.IntVar(value=0)
        self.ststk_type_var = tk.StringVar(value="Mat Kopya (50x70)")
        self.ststk_qty_var = tk.IntVar(value=1)
        self.nirengi_type_var = tk.StringVar(value="TUTGA noktası")
        self.nirengi_qty_var = tk.IntVar(value=1)
        self.parselasyon_type_var = tk.StringVar(value="Parselasyon Planı Kontrolü")
        self.previous_area_var = tk.StringVar(value="")
        self.aplikasyon_type_var = tk.StringVar(value="Normal aplikasyon / 2.1 yüzölçümü tarifesi")
        self.aplikasyon_parsel_turu_var = tk.StringVar(value="İmar Parseli")
        self.aplikasyon_contract_var = tk.StringVar(value="")
        self.aplikasyon_hectare_var = tk.StringVar(value="")
        self.aplikasyon_point_count_var = tk.IntVar(value=1)
        self.irtifak_type_var = tk.StringVar(value="Standart irtifak tesisi/terkini (2.4)")
        self.irtifak_include_docs_var = tk.BooleanVar(value=False)
        self.irtifak_corner_qty_var = tk.IntVar(value=0)
        self.irtifak_poligon_qty_var = tk.IntVar(value=0)
        self.irtifak_olcu_kroki_qty_var = tk.IntVar(value=0)
        self.irtifak_aplikasyon_kroki_qty_var = tk.IntVar(value=0)

        self.control_rows: dict[str, ttk.Frame] = {}
        self.control_row_labels: dict[str, ttk.Label] = {}
        self.parsel_turu_box = self.make_control_row(
            form,
            "parsel_turu",
            "Parsel Türü:",
            lambda parent: ttk.Combobox(
                parent,
                textvariable=self.parsel_turu_var,
                state="readonly",
                width=32,
                values=["Oluşmuş imar parselleri", "Kadastro parselleri"],
            ),
        )
        self.plan_parsel_turu_box = self.make_control_row(
            form,
            "plan_parsel_turu",
            "Plan/Kroki Parsel Türü:",
            lambda parent: ttk.Combobox(
                parent,
                textvariable=self.plan_parsel_turu_var,
                state="readonly",
                width=32,
                values=["İmar Parseli", "Kadastro Parseli", "Belediye Dışı (8/ğ)"],
            ),
        )
        self.cins_parsel_turu_box = self.make_control_row(
            form,
            "cins_parsel_turu",
            "Cins Değişikliği Parsel Türü:",
            lambda parent: ttk.Combobox(
                parent,
                textvariable=self.cins_parsel_turu_var,
                state="readonly",
                width=32,
                values=["İmar Parseli", "Kadastro Parseli"],
            ),
        )
        self.cins_type_box = self.make_control_row(
            form,
            "cins_type",
            "Cins değişikliği işlem türü:",
            lambda parent: ttk.Combobox(
                parent,
                textvariable=self.cins_type_var,
                state="readonly",
                width=56,
                values=[
                    "Yapısız iken yapılı hâle getirme (2.2.1)",
                    "Yapılı iken yapısız hâle gelme (2.2.2)",
                    "Yapı ile ilgisi olmayan vasıf değişikliği (2.2.3)",
                    "Aynı yapı üzerinde kat ilavesi (2.2.1.4.4)",
                    "Birden fazla yapı varsa kat ilavesi (2.2.1.4.5)",
                    "Aynı parselde sonradan yapılan yeni bina (2.2.1.4.8)",
                    "Yaygın kat mülkiyeti - arsa payına göre kısmi hesap (2.2.1.4.6)",
                ],
            ),
        )
        self.area_entry = self.make_control_row(form, "area", "Yüzölçümü (m²):", lambda parent: ttk.Entry(parent, textvariable=self.area_var, width=24))
        self.make_control_row(form, "km", "Uzunluk (km):", lambda parent: ttk.Entry(parent, textvariable=self.km_var, width=24))
        self.parselasyon_type_box = self.make_control_row(
            form,
            "parselasyon_type",
            "Kontrollük işlem türü:",
            lambda parent: ttk.Combobox(
                parent,
                textvariable=self.parselasyon_type_var,
                state="readonly",
                width=48,
                values=[
                    "Parselasyon Planı Kontrolü",
                    "Arazi Toplulaştırma Parselasyon Kontrolü",
                ],
            ),
        )
        self.make_control_row(
            form,
            "previous_area",
            "Daha önce aynı sözleşme kapsamında\nkontrolü/tahsilatı yapılan alan (m²):",
            lambda parent: ttk.Entry(parent, textvariable=self.previous_area_var, width=24),
        )
        self.make_control_row(
            form,
            "parsel_count",
            "Parsel sayısı:",
            lambda parent: ttk.Spinbox(parent, from_=1, to=10000, textvariable=self.parsel_count_var, width=12),
        )
        self.make_control_row(
            form,
            "cins_tarimsal",
            "Tarımsal amaçlı bina:",
            lambda parent: ttk.Checkbutton(parent, text="12.762 TL tavan uygula", variable=self.cins_tarimsal_var),
        )
        self.make_control_row(
            form,
            "cins_fazla_yapi",
            "Fazla yapı sayısı:",
            lambda parent: ttk.Spinbox(parent, from_=0, to=10000, textvariable=self.cins_fazla_yapi_var, width=12),
        )
        self.make_control_row(
            form,
            "cins_parsel_sayisi",
            "Parsel sayısı:",
            lambda parent: ttk.Spinbox(parent, from_=1, to=10000, textvariable=self.cins_parsel_sayisi_var, width=12),
        )
        self.make_control_row(
            form,
            "cins_yapi_sayisi",
            "Yapı / bina sayısı:",
            lambda parent: ttk.Spinbox(parent, from_=1, to=10000, textvariable=self.cins_yapi_sayisi_var, width=12),
        )
        self.make_control_row(
            form,
            "cins_arsa_pay_orani",
            "Arsa payı oranı (%):",
            lambda parent: ttk.Entry(parent, textvariable=self.cins_arsa_pay_orani_var, width=24),
        )
        self.kam_subtype_box = self.make_control_row(
            form,
            "kam_subtype",
            "Kamulaştırma türü:",
            lambda parent: ttk.Combobox(
                parent,
                textvariable=self.kam_subtype_var,
                state="readonly",
                width=28,
                values=["Şeritvari bazlı", "Hektar bazlı", "Hektar bazlı - MERA"],
            ),
        )
        self.ststk_type_box = self.make_control_row(
            form,
            "ststk_type",
            "ST/STK Türü:",
            lambda parent: ttk.Combobox(
                parent,
                textvariable=self.ststk_type_var,
                state="readonly",
                width=42,
                values=[
                    "Mat Kopya (50x70)",
                    "Mat Kopya (70x100)",
                    "Şeffaf Kopya (50x70)",
                    "Şeffaf Kopya (70x100)",
                    "Sayısal Harita Standart Pafta",
                    "Sayısal Arazi Modeli (20x20)",
                ],
            ),
        )
        self.make_control_row(
            form,
            "ststk_qty",
            "ST/STK Adet:",
            lambda parent: ttk.Spinbox(parent, from_=1, to=10000, textvariable=self.ststk_qty_var, width=12),
        )
        self.nirengi_type_box = self.make_control_row(
            form,
            "nirengi_type",
            "Nirengi Türü:",
            lambda parent: ttk.Combobox(
                parent,
                textvariable=self.nirengi_type_var,
                state="readonly",
                width=42,
                values=[
                    "TUTGA noktası",
                    "C1 derece AGA noktası",
                    "C2 derece SGA noktası",
                    "C3 derece ASN noktası",
                    "C4 derece poligon/fotogrametrik nokta",
                    "Ana nirengi noktası (ED-50)",
                    "Ara nirengi noktası (ED-50)",
                ],
            ),
        )
        self.make_control_row(
            form,
            "nirengi_qty",
            "Nirengi Adet:",
            lambda parent: ttk.Spinbox(parent, from_=1, to=10000, textvariable=self.nirengi_qty_var, width=12),
        )
        self.aplikasyon_type_box = self.make_control_row(
            form,
            "aplikasyon_type",
            "Aplikasyon işlem türü:",
            lambda parent: ttk.Combobox(
                parent,
                textvariable=self.aplikasyon_type_var,
                state="readonly",
                width=62,
                values=[
                    "Normal aplikasyon / 2.1 yüzölçümü tarifesi",
                    "Kimlik bilgisi güncelleme için zeminde inceleme / 1.10",
                    "Mera ihale aplikasyon kontrolü / 2.8.2",
                    "Mera GM hizmet talebi - aplikasyon yapımı / 2.8.3.2",
                    "Mera Komisyonu aplikasyon kontrolü / 2.8.4.1",
                    "Kadastral yol sınırlarının belirlenmesi / 2.9",
                ],
            ),
        )
        self.aplikasyon_parsel_turu_box = self.make_control_row(
            form,
            "aplikasyon_parsel_turu",
            "Aplikasyon parsel türü:",
            lambda parent: ttk.Combobox(
                parent,
                textvariable=self.aplikasyon_parsel_turu_var,
                state="readonly",
                width=34,
                values=["İmar Parseli", "Kadastro Parseli", "Belediye Dışı / 8/ğ"],
            ),
        )
        self.make_control_row(
            form,
            "aplikasyon_contract",
            "Sözleşme bedeli:",
            lambda parent: ttk.Entry(parent, textvariable=self.aplikasyon_contract_var, width=24),
        )
        self.make_control_row(
            form,
            "aplikasyon_hectare",
            "Hektar (ha):",
            lambda parent: ttk.Entry(parent, textvariable=self.aplikasyon_hectare_var, width=24),
        )
        self.make_control_row(
            form,
            "aplikasyon_point_count",
            "Nokta sayısı:",
            lambda parent: ttk.Spinbox(parent, from_=1, to=100000, textvariable=self.aplikasyon_point_count_var, width=12),
        )
        self.report_type_box = self.make_control_row(
            form,
            "report_type",
            "Mülkiyet raporu türü:",
            lambda parent: ttk.Combobox(
                parent,
                textvariable=self.report_type_var,
                state="readonly",
                width=38,
                values=[
                    "İmar ayırma/birleştirme (parsel)",
                    "Parselasyon / toplulaştırma / hektar bazlı kamulaştırma (ha)",
                    "Şeritvari kamulaştırma (km)",
                ],
            ),
        )
        self.make_control_row(
            form,
            "report_parsel",
            "Mülkiyet raporu parsel adedi:",
            lambda parent: ttk.Spinbox(parent, from_=1, to=10000, textvariable=self.report_parsel_var, width=12),
        )
        self.irtifak_type_box = self.make_control_row(
            form,
            "irtifak_type",
            "İrtifak işlem türü:",
            lambda parent: ttk.Combobox(
                parent,
                textvariable=self.irtifak_type_var,
                state="readonly",
                width=58,
                values=[
                    "Standart irtifak tesisi/terkini (2.4)",
                    "Kamulaştırma Kanunu kapsamında hektar bazlı irtifak/terkin (2.11.5.2.4)",
                    "Kamulaştırma yoluyla irtifak hakkı tesisi - hektar bazlı kontrol (2.11.5.2)",
                    "Şeritvari bazlı kamulaştırma/irtifak kontrolü (2.11.5.1)",
                ],
            ),
        )
        self.irtifak_docs_check = self.make_control_row(
            form,
            "irtifak_docs",
            "Harita bilgi-belge eklensin:",
            lambda parent: ttk.Checkbutton(parent, text="Evet", variable=self.irtifak_include_docs_var),
        )
        self.make_control_row(
            form,
            "irtifak_corner_qty",
            "Parsel köşe koordinatı nokta sayısı:",
            lambda parent: ttk.Spinbox(parent, from_=0, to=500000, textvariable=self.irtifak_corner_qty_var, width=12),
        )
        self.make_control_row(
            form,
            "irtifak_poligon_qty",
            "Poligon koordinatı adedi:",
            lambda parent: ttk.Spinbox(parent, from_=0, to=500000, textvariable=self.irtifak_poligon_qty_var, width=12),
        )
        self.make_control_row(
            form,
            "irtifak_olcu_kroki_qty",
            "Ölçü krokisi sayfa sayısı:",
            lambda parent: ttk.Spinbox(parent, from_=0, to=500000, textvariable=self.irtifak_olcu_kroki_qty_var, width=12),
        )
        self.make_control_row(
            form,
            "irtifak_aplikasyon_kroki_qty",
            "Aplikasyon krokisi sayfa sayısı:",
            lambda parent: ttk.Spinbox(parent, from_=0, to=500000, textvariable=self.irtifak_aplikasyon_kroki_qty_var, width=12),
        )
        self.make_control_row(form, "geriye", "Geriye dönüş işlemi:", lambda parent: ttk.Checkbutton(parent, variable=self.geriye_var))
        self.make_control_row(
            form,
            "yeni",
            "Geriye dönüş + yeni uygulama:",
            lambda parent: ttk.Checkbutton(parent, variable=self.yeni_var),
        )
        self.mera_check = self.make_control_row(
            form,
            "mera",
            "Mera Kanunu kapsamı (%25):",
            lambda parent: ttk.Checkbutton(parent, variable=self.mera_var),
        )
        self.make_control_row(
            form,
            "docs",
            "Harita bilgi-belge dahil:",
            lambda parent: ttk.Checkbutton(parent, variable=self.include_docs_var),
        )
        self.make_control_row(
            form,
            "hatali_bagimsiz_sayi",
            "Hatalı bağımsız bölüm adedi:",
            lambda parent: ttk.Spinbox(parent, from_=0, to=10000, textvariable=self.hatali_bagimsiz_sayi_var, width=12),
        )
        self.make_control_row(
            form,
            "hatali_blok_sayi",
            "Hatalı blok adedi:",
            lambda parent: ttk.Spinbox(parent, from_=0, to=10000, textvariable=self.hatali_blok_sayi_var, width=12),
        )

        self.kam_subtype_box.bind("<<ComboboxSelected>>", self.on_control_type_change)
        self.report_type_box.bind("<<ComboboxSelected>>", self.on_control_type_change)
        self.parsel_turu_box.bind("<<ComboboxSelected>>", self.on_control_type_change)
        self.cins_type_box.bind("<<ComboboxSelected>>", self.on_control_type_change)
        self.aplikasyon_type_box.bind("<<ComboboxSelected>>", self.on_control_type_change)
        self.irtifak_type_box.bind("<<ComboboxSelected>>", self.on_control_type_change)
        self.parselasyon_type_box.bind("<<ComboboxSelected>>", self.on_control_type_change)
        self.irtifak_docs_check.configure(command=self.on_control_type_change)
        area_validator = (self.register(self.validate_area_input), "%P")
        self.area_entry.configure(validate="key", validatecommand=area_validator)

        action = ttk.Frame(right)
        action.pack(fill="x", pady=(0, 8))
        ttk.Button(action, text="Hesapla", style="Primary.TButton", command=self.calculate_control).pack(side="left")
        self.control_total_label = ttk.Label(action, text="Kontrollük Ücreti: -", style="Info.TLabel")
        self.control_total_label.pack(side="left", padx=(12, 0))

        result_box = ttk.LabelFrame(right, text="Kontrollük Sonucu", padding=8)
        result_box.pack(fill="both", expand=True)
        self.control_result = tk.Text(result_box, height=16, wrap="word", relief="flat", borderwidth=0, padx=8, pady=8)
        self.control_result.pack(fill="both", expand=True)
        self.theme_targets.append(self.control_result)
        self.control_result.insert("1.0", "Seçili işlem için parametreleri girip Hesapla butonuna basın.")

        self.refresh_control_iller()
        self.set_active_operation("degisiklik")

    @staticmethod
    def validate_area_input(proposed: str) -> bool:
        if proposed == "":
            return True
        return bool(re.fullmatch(r"\d+(?:[.,]\d{0,2})?", proposed))

    def make_control_row(
        self, parent: ttk.Frame, key: str, label_text: str, widget_factory: Callable[[ttk.Frame], tk.Widget]
    ) -> tk.Widget:
        row = len(self.control_rows)
        frame = ttk.Frame(parent)
        frame.grid(row=row, column=0, sticky="w", pady=3)
        label = ttk.Label(frame, text=label_text, width=34)
        if key == "previous_area":
            label.configure(wraplength=280, justify="left")
        label.pack(side="left")
        widget = widget_factory(frame)
        widget.pack(side="left")
        self.control_rows[key] = frame
        self.control_row_labels[key] = label
        return widget

    def refresh_control_iller(self) -> None:
        iller = sorted(self.il_index.keys())
        if not iller:
            self.control_il_box.configure(values=[])
            self.control_ilce_box.configure(values=[])
            return
        self.control_il_box.configure(values=iller)
        default_il = normalize_key(self.settings.get("default_il", iller[0]))
        if default_il not in iller:
            default_il = iller[0]
        self.control_il_var.set(default_il)
        self.on_control_il_change()

    def on_control_il_change(self, _event: object | None = None) -> None:
        il = self.control_il_var.get()
        ilceler = sorted(self.il_index.get(il, {}).keys())
        self.control_ilce_box.configure(values=ilceler)
        if ilceler:
            if self.control_ilce_var.get() not in ilceler:
                self.control_ilce_var.set(ilceler[0])
        else:
            self.control_ilce_var.set("")
        self.update_control_yk()

    def update_control_yk(self) -> None:
        il = self.control_il_var.get()
        ilce = self.control_ilce_var.get()
        yk = self.il_index.get(il, {}).get(ilce, {}).get("kadastro")
        self.yk_label.configure(text=f"Kadastro YK: {yk:g}" if isinstance(yk, float) else "Kadastro YK: -")
        if hasattr(self, "mudurluk_preview_label"):
            self.mudurluk_preview_label.configure(text=f"Müdürlük: {mudurluk_name_from_il(il)}")

    def set_active_operation(self, operation_code: str) -> None:
        self.operation_code = operation_code
        title_map = {code: title for title, code in self.operation_defs}
        self.selected_process_label.configure(text=title_map.get(operation_code, "-"))

        for code, btn in self.operation_buttons.items():
            btn.configure(style="OperationActive.TButton" if code == operation_code else "Operation.TButton")

        if operation_code == "degisiklik":
            self.control_type_var.set("degisiklik")
        elif operation_code == "parselasyon":
            self.control_type_var.set("parselasyon")
        elif operation_code == "kamulastirma":
            self.control_type_var.set("kamulastirma")
        elif operation_code == "mulkiyet":
            self.control_type_var.set("mulkiyet")
        elif operation_code == "irtifak":
            self.control_type_var.set("irtifak")
        elif operation_code == "birlestirme":
            self.control_type_var.set("birlestirme")
        elif operation_code == "cins":
            self.control_type_var.set("cins")
        elif operation_code == "plan_kroki":
            self.control_type_var.set("plan_kroki")
        elif operation_code == "aplikasyon":
            self.control_type_var.set("aplikasyon")
        elif operation_code == "parsel_yerinde":
            self.control_type_var.set("parsel_yerinde")
        elif operation_code == "hatali_bagimsiz":
            self.control_type_var.set("hatali_bagimsiz")
        elif operation_code == "st_stk":
            self.control_type_var.set("st_stk")
        elif operation_code == "nirengi":
            self.control_type_var.set("nirengi")
        else:
            self.control_type_var.set(operation_code)

        self.on_control_type_change()
        if hasattr(self, "control_result"):
            self.clear_control_result()

    def clear_control_result(self) -> None:
        self.control_total_label.configure(text="Kontrollük Ücreti: -")
        self.control_result.delete("1.0", "end")
        self.control_result.insert("1.0", "Seçili işlem için parametreleri girip Hesapla butonuna basın.")

    def on_control_type_change(self, _event: object | None = None) -> None:
        show = set()
        self.control_row_labels["area"].configure(text="Yüzölçümü (m²):")
        self.control_row_labels["parsel_count"].configure(text="Parsel sayısı:")
        if self.operation_code == "parselasyon":
            show = {"parselasyon_type", "area", "geriye", "yeni"}
            self.control_row_labels["area"].configure(text="Yüzölçümü (m²):")
            if self.parselasyon_type_var.get().startswith("Arazi Toplulaştırma"):
                show.add("previous_area")
                show.discard("geriye")
                show.discard("yeni")
                self.geriye_var.set(False)
                self.yeni_var.set(False)
                self.control_row_labels["area"].configure(text="Bu kontrole gelen alan (m²):")
        elif self.operation_code == "degisiklik":
            show = {"parsel_turu", "area", "parsel_count", "mera"}
        elif self.operation_code == "kamulastirma":
            show = {"kam_subtype", "docs"}
            if self.kam_subtype_var.get() == "Şeritvari bazlı":
                show.add("km")
            else:
                show.add("area")
                self.control_row_labels["area"].configure(text="Hektar (ha):")
        elif self.operation_code == "mulkiyet":
            show = {"report_type"}
            report_type = self.report_type_var.get()
            if report_type.startswith("İmar ayırma"):
                show.add("report_parsel")
            elif report_type.startswith("Parselasyon") or "hektar" in tr_lower(report_type):
                show.add("area")
                self.control_row_labels["area"].configure(text="Hektar (ha):")
            else:
                show.add("km")
        elif self.operation_code == "irtifak":
            show = {"irtifak_type", "parsel_count"}
            self.control_row_labels["parsel_count"].configure(text="İşleme konu parsel sayısı:")
            irtifak_type = self.irtifak_type_var.get()
            if "2.11.5.2.4" in irtifak_type:
                show.add("irtifak_docs")
                if self.irtifak_include_docs_var.get():
                    show.update(
                        {
                            "irtifak_corner_qty",
                            "irtifak_poligon_qty",
                            "irtifak_olcu_kroki_qty",
                            "irtifak_aplikasyon_kroki_qty",
                        }
                    )
            elif "hektar bazlı kontrol" in tr_lower(irtifak_type):
                show.update({"area", "docs"})
                self.control_row_labels["area"].configure(text="Hektar (ha):")
            elif "şeritvari" in tr_lower(irtifak_type):
                show.update({"km", "docs"})
        elif self.operation_code in {"birlestirme", "parsel_yerinde"}:
            show = {"parsel_count"}
        elif self.operation_code == "aplikasyon":
            app_type = self.aplikasyon_type_var.get()
            show = {"aplikasyon_type"}
            if app_type.startswith("Normal aplikasyon"):
                show.update({"aplikasyon_parsel_turu", "area"})
                self.control_row_labels["area"].configure(text="Yüzölçümü (m²):")
            elif app_type.startswith("Kimlik bilgisi"):
                show.update({"parsel_count"})
                self.control_row_labels["parsel_count"].configure(text="Taşınmaz sayısı:")
            elif app_type.startswith("Mera ihale aplikasyon kontrolü"):
                show.update({"aplikasyon_contract"})
            elif app_type.startswith("Mera GM hizmet talebi"):
                show.update({"aplikasyon_hectare"})
            elif app_type.startswith("Mera Komisyonu aplikasyon kontrolü"):
                show.update({"aplikasyon_hectare"})
            elif app_type.startswith("Kadastral yol"):
                show.update({"aplikasyon_point_count"})
        elif self.operation_code == "cins":
            cins_type = self.cins_type_var.get()
            show = {"cins_type"}
            if cins_type.startswith("Yapısız"):
                show.update({"cins_parsel_turu", "area", "cins_tarimsal", "cins_fazla_yapi"})
            elif cins_type.startswith("Yapılı iken yapısız"):
                show.update({"cins_parsel_sayisi"})
            elif cins_type.startswith("Yapı ile ilgisi olmayan"):
                show.update({"cins_parsel_sayisi"})
            elif cins_type.startswith("Aynı yapı üzerinde kat ilavesi"):
                show.update({"cins_yapi_sayisi"})
            elif cins_type.startswith("Birden fazla yapı varsa kat ilavesi"):
                show.update({"cins_yapi_sayisi"})
            elif cins_type.startswith("Aynı parselde sonradan"):
                show.update({"cins_yapi_sayisi"})
            elif cins_type.startswith("Yaygın kat mülkiyeti"):
                show.update({"cins_parsel_turu", "area", "cins_arsa_pay_orani"})
        elif self.operation_code == "plan_kroki":
            show = {"area", "plan_parsel_turu"}
        elif self.operation_code == "hatali_bagimsiz":
            show = {"hatali_bagimsiz_sayi", "hatali_blok_sayi"}
        elif self.operation_code == "st_stk":
            show = {"ststk_type", "ststk_qty"}
        elif self.operation_code == "nirengi":
            show = {"nirengi_type", "nirengi_qty"}
        for key, frame in self.control_rows.items():
            if key in show:
                frame.grid()
            else:
                frame.grid_remove()

        if self.operation_code == "degisiklik" and "Oluşmuş imar" in self.parsel_turu_var.get():
            self.mera_var.set(False)
            self.mera_check.configure(state="disabled")
        else:
            self.mera_check.configure(state="normal")

    def get_selected_yk(self) -> float:
        il = self.control_il_var.get()
        ilce = self.control_ilce_var.get()
        info = self.il_index.get(il, {}).get(ilce, {})
        return float(info.get("kadastro", 1.0))

    def calculate_control(self) -> None:
        calc = self.operation_code
        yk = self.get_selected_yk()
        rate_map = self.control_rates
        try:
            area_value = max(parse_decimal_input(self.area_var.get(), 0.0), 0.0)
            km_value = max(parse_decimal_input(self.km_var.get(), 0.0), 0.0)
        except ValueError:
            messagebox.showerror("Hata", "Yüzölçümü/uzunluk alanında geçersiz sayı girdiniz.")
            return

        details: list[str] = []
        kam_breakdown: dict[str, float | bool | str] | None = None
        raw = 0.0
        uses_yk = True
        if calc == "parselasyon":
            geriye_flag = self.geriye_var.get() or self.yeni_var.get()
            yeni_flag = self.yeni_var.get()
            try:
                previous_area_value = max(parse_decimal_input(self.previous_area_var.get(), 0.0), 0.0)
            except ValueError:
                messagebox.showerror("Hata", "Önceki alan değerinde geçersiz sayı girdiniz.")
                return
            if self.parselasyon_type_var.get().startswith("Arazi Toplulaştırma"):
                raw, details = toplulastirma_ilave_basamak_fee(
                    current_area_m2=area_value,
                    previous_area_m2=previous_area_value,
                    yk=yk,
                    geriye_donus=geriye_flag,
                    yeni_uygulama=yeni_flag,
                    rate_map=rate_map,
                )
            else:
                raw, details = parselasyon_fee(
                    area_m2=area_value,
                    yk=yk,
                    geriye_donus=geriye_flag,
                    yeni_uygulama=yeni_flag,
                    rate_map=rate_map,
                )
        elif calc == "degisiklik" and "Oluşmuş" in self.parsel_turu_var.get():
            effective_parsel_count = max(int(self.parsel_count_var.get() or 0), 2)
            if effective_parsel_count != int(self.parsel_count_var.get() or 0):
                details.append("İfraz/tevhit işlemlerinde minimum parsel sayısı 2 olarak alındı.")
                self.parsel_count_var.set(2)
            raw, fee_details = imar_degisiklik_fee(
                area_value, yk, effective_parsel_count, self.mera_var.get(), rate_map
            )
            details.extend(fee_details)
        elif calc == "degisiklik":
            effective_parsel_count = max(int(self.parsel_count_var.get() or 0), 2)
            if effective_parsel_count != int(self.parsel_count_var.get() or 0):
                details.append("İfraz/tevhit işlemlerinde minimum parsel sayısı 2 olarak alındı.")
                self.parsel_count_var.set(2)
            raw, fee_details = kadastro_degisiklik_fee(
                area_value, yk, effective_parsel_count, self.mera_var.get(), rate_map
            )
            details.extend(fee_details)
        elif calc == "kamulastirma":
            subtype = self.kam_subtype_var.get()
            if subtype == "Şeritvari bazlı":
                eff_km = max(float(km_value or 0), 1.0)
                serit_kontrol = self.get_rate("kam_serit_kontrol")
                serit_docs = self.get_rate("kam_serit_docs")
                control = eff_km * serit_kontrol
                docs = eff_km * serit_docs if self.include_docs_var.get() else 0.0
                raw = control + docs
                details = [f"Şeritvari kontrol: {eff_km:g} km x {tr_money(serit_kontrol)} = {tr_money(control)}"]
                if self.include_docs_var.get():
                    details.append(f"Harita bilgi ve belge: {eff_km:g} km x {tr_money(serit_docs)} = {tr_money(docs)}")
                kam_breakdown = {
                    "control": control,
                    "docs": docs,
                    "include_docs": self.include_docs_var.get(),
                    "subtype": subtype,
                }
                uses_yk = False
            else:
                raw_ha = max(float(area_value or 0), 1.0)
                control_ha = min(raw_ha, 100.0)
                docs_ha = raw_ha
                hektar_kontrol = self.get_rate("kam_hektar_kontrol")
                hektar_docs = self.get_rate("kam_hektar_docs")
                control = control_ha * hektar_kontrol
                docs = docs_ha * hektar_docs if self.include_docs_var.get() else 0.0
                mera_25 = subtype.endswith("MERA")
                mera_factor = 0.25 if mera_25 else 1.0
                control_adj = control * mera_factor * yk
                docs_adj = docs * mera_factor * yk
                raw = control_adj + docs_adj
                details = [f"Hektar bazlı kontrol: {control_ha:g} ha x {tr_money(hektar_kontrol)} = {tr_money(control)}"]
                if raw_ha > 100.0:
                    details.append("Kontrollük kısmında 100 ha üstü için 100 ha sınırı uygulandı.")
                if self.include_docs_var.get():
                    details.append(f"Harita bilgi ve belge: {docs_ha:g} ha x {tr_money(hektar_docs)} = {tr_money(docs)}")
                if mera_25:
                    details.append("Mera Kanunu özel durumu: %25 tahsil")
                details.append(f"Yöresel katsayı uygulandı: {yk:g}")
                kam_breakdown = {
                    "control": control_adj,
                    "docs": docs_adj,
                    "include_docs": self.include_docs_var.get(),
                    "subtype": subtype,
                    "mera_25": mera_25,
                    "yk": yk,
                }
        elif calc == "mulkiyet":
            raw, details = mulkiyet_raporu_fee(
                self.report_type_var.get(), area_value, km_value, self.report_parsel_var.get(), rate_map
            )
            uses_yk = False
        elif calc == "irtifak":
            raw, details, uses_yk = irtifak_fee(
                parsel_count=self.parsel_count_var.get(),
                irtifak_type=self.irtifak_type_var.get(),
                yk=yk,
                hektar=area_value,
                km=km_value,
                include_docs=(
                    self.irtifak_include_docs_var.get()
                    if "2.11.5.2.4" in self.irtifak_type_var.get()
                    else self.include_docs_var.get()
                ),
                corner_qty=self.irtifak_corner_qty_var.get(),
                poligon_qty=self.irtifak_poligon_qty_var.get(),
                olcu_kroki_qty=self.irtifak_olcu_kroki_qty_var.get(),
                aplikasyon_kroki_qty=self.irtifak_aplikasyon_kroki_qty_var.get(),
                rate_map=rate_map,
            )
        elif calc == "birlestirme":
            uses_yk = False
            count = max(int(self.parsel_count_var.get() or 0), 2)
            if count != int(self.parsel_count_var.get() or 0):
                self.parsel_count_var.set(2)
                details.append("Tevhit işleminde minimum parsel sayısı 2 olarak alındı.")
            birlestirme_birim = 2227.0
            raw = max(count - 1, 1) * birlestirme_birim
            details.append(f"((n-1) x 2.227 TL) formülü: ({count}-1) x {tr_money(birlestirme_birim)} = {tr_money(raw)}")
        elif calc == "cins":
            try:
                arsa_pay_orani = parse_decimal_input(self.cins_arsa_pay_orani_var.get(), 100.0)
            except ValueError:
                messagebox.showerror("Hata", "Arsa payı oranı alanında geçersiz sayı girdiniz.")
                return
            raw, details, uses_yk = cins_degisikligi_detayli_fee(
                cins_type=self.cins_type_var.get(),
                area_m2=area_value,
                yk=yk,
                parsel_turu=self.cins_parsel_turu_var.get(),
                tarimsal_tavan=self.cins_tarimsal_var.get(),
                fazla_yapi_sayisi=self.cins_fazla_yapi_var.get(),
                parsel_sayisi=self.cins_parsel_sayisi_var.get(),
                yapi_sayisi=self.cins_yapi_sayisi_var.get(),
                arsa_pay_orani=arsa_pay_orani,
                rate_map=rate_map,
            )
        elif calc == "plan_kroki":
            raw, details = plan_kroki_fee(area_value, yk, self.plan_parsel_turu_var.get(), rate_map)
        elif calc == "aplikasyon":
            try:
                app_contract = parse_decimal_input(self.aplikasyon_contract_var.get(), 0.0)
                app_hectare = parse_decimal_input(self.aplikasyon_hectare_var.get(), 0.0)
            except ValueError:
                messagebox.showerror("Hata", "Aplikasyon alanlarında geçersiz sayı girdiniz.")
                return
            raw, details, uses_yk = aplikasyon_fee(
                aplikasyon_type=self.aplikasyon_type_var.get(),
                area_m2=area_value,
                yk=yk,
                parsel_turu=self.aplikasyon_parsel_turu_var.get(),
                parsel_count=self.parsel_count_var.get(),
                contract_amount=app_contract,
                hectare=app_hectare,
                point_count=self.aplikasyon_point_count_var.get(),
                rate_map=rate_map,
            )
        elif calc == "parsel_yerinde":
            uses_yk = False
            count = max(int(self.parsel_count_var.get() or 0), 1)
            yerinde_ilk = self.get_rate("cins_yerinde_gosterim")
            yerinde_ilave = self.get_rate("kam_hektar_docs")
            raw = yerinde_ilk + max(count - 1, 0) * yerinde_ilave
            details = [
                f"İlk parsel: {tr_money(yerinde_ilk)}",
                f"İlave {max(count-1,0)} parsel x {tr_money(yerinde_ilave)} = {tr_money(max(count-1,0)*yerinde_ilave)}",
            ]
        elif calc == "hatali_bagimsiz":
            uses_yk = False
            bagimsiz = max(int(self.hatali_bagimsiz_sayi_var.get() or 0), 0)
            blok = max(int(self.hatali_blok_sayi_var.get() or 0), 0)
            bagimsiz_total = 0.0
            if bagimsiz > 0:
                birim_hata = 2227.0
                ilave_hata = self.get_rate("kam_hektar_docs")
                bagimsiz_total = birim_hata if bagimsiz <= 2 else birim_hata + (bagimsiz - 2) * ilave_hata
            blok_birim = 2227.0
            blok_total = blok * blok_birim
            raw = bagimsiz_total + blok_total
            details = [
                f"Bağımsız bölüm düzeltme bedeli: {tr_money(bagimsiz_total)}",
                f"Blok düzeltme bedeli: {blok} x {tr_money(blok_birim)} = {tr_money(blok_total)}",
            ]
        elif calc == "st_stk":
            uses_yk = False
            prices = {
                "Mat Kopya (50x70)": 518.0,
                "Mat Kopya (70x100)": 720.0,
                "Şeffaf Kopya (50x70)": 720.0,
                "Şeffaf Kopya (70x100)": 955.0,
                "Sayısal Harita Standart Pafta": 6686.0,
                "Sayısal Arazi Modeli (20x20)": 240.0,
            }
            sel = self.ststk_type_var.get()
            price = prices.get(sel, 0.0)
            qty = max(int(self.ststk_qty_var.get() or 0), 1)
            raw = price * qty
            details = [f"{sel}: {qty} x {tr_money(price)} = {tr_money(raw)}"]
        elif calc == "nirengi":
            uses_yk = False
            prices = {
                "TUTGA noktası": 1773.0,
                "C1 derece AGA noktası": 1773.0,
                "C2 derece SGA noktası": 1033.0,
                "C3 derece ASN noktası": 778.0,
                "C4 derece poligon/fotogrametrik nokta": 338.0,
                "Ana nirengi noktası (ED-50)": 665.0,
                "Ara nirengi noktası (ED-50)": 338.0,
            }
            sel = self.nirengi_type_var.get()
            price = prices.get(sel, 0.0)
            qty = max(int(self.nirengi_qty_var.get() or 0), 1)
            raw = price * qty
            details = [f"{sel}: {qty} x {tr_money(price)} = {tr_money(raw)}"]

        min_kontrolluk = self.get_rate("min_kontrolluk")
        kontrolluk_ops = {"parselasyon", "degisiklik", "kamulastirma"}
        final = max(raw, min_kontrolluk) if calc in kontrolluk_ops else raw
        if calc == "kamulastirma" and kam_breakdown is not None:
            control_fee = float(kam_breakdown.get("control", 0.0))
            docs_fee = float(kam_breakdown.get("docs", 0.0))
            lines = [f"Kontrollük Hizmet: {tr_money(final)}", ""]
            lines.append(f"└─ Kontrollük: {tr_money(control_fee)}")
            if bool(kam_breakdown.get("include_docs", False)):
                lines.append(f"└─ Teknik Bilgi-Belge: {tr_money(docs_fee)}")
            if bool(kam_breakdown.get("mera_25", False)):
                lines.append("└─ Mera Kanunu özel durumu: %25 tahsil")
            if "yk" in kam_breakdown:
                lines.append(f"└─ Yöresel katsayı: {float(kam_breakdown['yk']):g}")
            if final > raw:
                lines.append(f"└─ Asgari kontrollük sınırı uygulandı: {tr_money(min_kontrolluk)}")
            lines.append("")
            lines.append("Detay:")
            lines.extend(f"- {line}" for line in details)
        else:
            lines = [
                f"Hesaplanan bedel: {tr_money(raw)}",
                f"Tahsil edilecek bedel: {tr_money(final)}",
            ]
            if calc in kontrolluk_ops and final > raw:
                lines.append(f"Asgari kontrollük sınırı uygulandı: {tr_money(min_kontrolluk)}")
            lines.append("")
            lines.append("Detay:")
            lines.extend(f"- {line}" for line in details)
            if uses_yk and calc not in {"parselasyon", "degisiklik", "kamulastirma", "mulkiyet", "cins", "plan_kroki"}:
                lines.append(f"- Yöresel katsayı: {yk:g}")
            elif uses_yk:
                lines.append(f"- Yöresel katsayı: {yk:g}")
            else:
                lines.append("- Yöresel katsayı uygulanmaz (maktu)")
        self.control_result.delete("1.0", "end")
        self.control_result.insert("1.0", "\n".join(lines))
        self.control_total_label.configure(text=f"Kontrollük Ücreti: {tr_money(final)}")

    def build_admin_tab(self) -> None:
        top = ttk.Frame(self.tab_admin)
        top.pack(fill="x", pady=(0, 8))
        ttk.Label(top, text="Varsayılan il:").pack(side="left", padx=(0, 6))
        self.admin_default_il_var = tk.StringVar()
        self.admin_default_il_box = ttk.Combobox(top, textvariable=self.admin_default_il_var, state="readonly", width=20)
        self.admin_default_il_box.pack(side="left")
        ttk.Button(top, text="Kaydet", command=self.save_default_il).pack(side="left", padx=8)
        ttk.Button(top, text="Excel'den Yükle", style="Primary.TButton", command=self.import_excel_coefficients).pack(side="left", padx=(12, 6))
        ttk.Button(top, text="Şablonu İndir", command=self.download_excel_template).pack(side="left")
        self.admin_year_label = ttk.Label(top, text=f"Aktif Yıl: {self.selected_year}", style="Info.TLabel")
        self.admin_year_label.pack(side="right")

        tabs = ttk.Notebook(self.tab_admin)
        tabs.pack(fill="both", expand=True)

        self.admin_coeff_tab = ttk.Frame(tabs, padding=10)
        self.admin_rates_tab = ttk.Frame(tabs, padding=10)
        tabs.add(self.admin_coeff_tab, text="Yöresel Katsayılar")
        tabs.add(self.admin_rates_tab, text="Kontrollük Ücreti Tutarları")

        list_frame = ttk.Frame(self.admin_coeff_tab)
        list_frame.pack(fill="both", expand=True)
        cols = ("bolge", "il", "ilce", "tapu", "kadastro")
        self.admin_tree = ttk.Treeview(list_frame, columns=cols, show="headings")
        for col, title, width in [
            ("bolge", "Bölge", 120),
            ("il", "İl", 120),
            ("ilce", "İlçe", 160),
            ("tapu", "Tapu YK", 90),
            ("kadastro", "Kadastro YK", 110),
        ]:
            self.admin_tree.heading(col, text=title)
            self.admin_tree.column(col, width=width, anchor="w")
        self.admin_tree.pack(side="left", fill="both", expand=True)
        self.admin_tree.bind("<Double-1>", self.on_admin_row_double_click)
        scroll = ttk.Scrollbar(list_frame, orient="vertical", command=self.admin_tree.yview)
        scroll.pack(side="left", fill="y")
        self.admin_tree.configure(yscrollcommand=scroll.set)

        rates_frame = ttk.Frame(self.admin_rates_tab)
        rates_frame.pack(fill="both", expand=True)
        self.rate_tree = ttk.Treeview(rates_frame, columns=("kod", "aciklama", "tutar"), show="headings")
        self.rate_tree.heading("kod", text="Kod")
        self.rate_tree.heading("aciklama", text="Açıklama")
        self.rate_tree.heading("tutar", text="Tutar")
        self.rate_tree.column("kod", width=210, anchor="w")
        self.rate_tree.column("aciklama", width=500, anchor="w")
        self.rate_tree.column("tutar", width=130, anchor="e")
        self.rate_tree.pack(side="left", fill="both", expand=True)
        self.rate_tree.bind("<Double-1>", self.on_rate_row_double_click)
        rate_scroll = ttk.Scrollbar(rates_frame, orient="vertical", command=self.rate_tree.yview)
        rate_scroll.pack(side="left", fill="y")
        self.rate_tree.configure(yscrollcommand=rate_scroll.set)

        ttk.Label(
            self.admin_coeff_tab,
            text="Not: Şablonu indirip yükledikten sonra manuel düzeltme için listede il-ilçe satırına çift tıklayarak güncelleyin.",
            style="Muted.TLabel",
        ).pack(anchor="w", pady=(8, 0))
        ttk.Label(
            self.admin_rates_tab,
            text="Not: Kontrollük tutarları satırına çift tıklayarak güncelleyebilir veya Excel ile toplu yükleyebilirsiniz.",
            style="Muted.TLabel",
        ).pack(anchor="w", pady=(8, 0))
        ttk.Label(self.tab_admin, text=f"Ayar dosyası: {SETTINGS_FILE}").pack(anchor="w", pady=(8, 0))
        self.refresh_admin_ui()

    def refresh_admin_ui(self) -> None:
        iller = sorted(self.il_index.keys())
        self.admin_default_il_box.configure(values=iller)
        selected = normalize_key(self.settings.get("default_il", iller[0] if iller else ""))
        if selected not in iller and iller:
            selected = iller[0]
        self.admin_default_il_var.set(selected)
        if hasattr(self, "year_box"):
            self.year_box.configure(values=[str(y) for y in self.get_year_options()])
        if hasattr(self, "admin_year_label"):
            self.admin_year_label.configure(text=f"Aktif Yıl: {self.selected_year}")

        for row_id in self.admin_tree.get_children():
            self.admin_tree.delete(row_id)
        for il in sorted(self.il_index.keys()):
            for ilce in sorted(self.il_index[il].keys()):
                info = self.il_index[il][ilce]
                self.admin_tree.insert(
                    "",
                    "end",
                    values=(info["bolge"], il, ilce, f'{info["tapu"]:g}', f'{info["kadastro"]:g}'),
                )
        self.refresh_rate_ui()
        self.refresh_mahkeme_infaz_label()

    def refresh_rate_ui(self) -> None:
        if not hasattr(self, "rate_tree"):
            return
        for row_id in self.rate_tree.get_children():
            self.rate_tree.delete(row_id)
        for code, label, default_value in CONTROL_RATE_DEFS:
            value = self.control_rates.get(code, default_value)
            self.rate_tree.insert("", "end", values=(code, label, f"{float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")))

    def save_default_il(self) -> None:
        selected = normalize_key(self.admin_default_il_var.get().strip())
        if not selected:
            messagebox.showerror("Hata", "Varsayılan il boş olamaz.")
            return
        self.settings["default_il"] = selected
        self.save_settings()
        self.refresh_control_iller()
        messagebox.showinfo("Bilgi", "Varsayılan il kaydedildi.")

    def download_excel_template(self) -> None:
        if Workbook is None:
            err = f"\nDetay: {OPENPYXL_IMPORT_ERROR}" if OPENPYXL_IMPORT_ERROR else ""
            messagebox.showerror("Hata", f"Excel şablonu için 'openpyxl' bulunamadı.{err}")
            return

        target = filedialog.asksaveasfilename(
            title="Excel Şablonu Kaydet",
            defaultextension=".xlsx",
            initialfile="Katsayi_Harc.xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not target:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Katsayilar"
        ws.append(EXCEL_HEADERS)
        for il in sorted(self.il_index.keys()):
            for ilce in sorted(self.il_index[il].keys()):
                info = self.il_index[il][ilce]
                ws.append([info["bolge"], il, ilce, float(info["tapu"]), float(info["kadastro"])])

        ws_rates = wb.create_sheet("KontrollukUcretleri")
        ws_rates.append(RATE_EXCEL_HEADERS)
        for code, label, default_value in CONTROL_RATE_DEFS:
            ws_rates.append([code, label, float(self.control_rates.get(code, default_value))])
        wb.save(target)
        messagebox.showinfo(
            "Bilgi",
            (
                f"Şablon kaydedildi:\n{target}\n\n"
                "İçerik: Katsayilar + KontrollukUcretleri.\n"
                "Manuel düzeltme için listelerde satıra çift tıklayabilirsiniz."
            ),
        )

    def import_excel_coefficients(self) -> None:
        if load_workbook is None:
            err = f"\nDetay: {OPENPYXL_IMPORT_ERROR}" if OPENPYXL_IMPORT_ERROR else ""
            messagebox.showerror("Hata", f"Excel içe aktarımı için 'openpyxl' bulunamadı.{err}")
            return

        selected = filedialog.askopenfilename(
            title="Yöresel Katsayı Excel Seç",
            filetypes=[("Excel", "*.xlsx"), ("Excel Macro", "*.xlsm")],
        )
        if not selected:
            return

        try:
            wb = load_workbook(selected, data_only=True)

            ws_coeff = wb["Katsayilar"] if "Katsayilar" in wb.sheetnames else wb.active
            header_row = [str(cell.value or "").strip() for cell in ws_coeff[1]]
            normalized = [canonical_header(v) for v in header_row]
            expected = [canonical_header(h) for h in EXCEL_HEADERS]
            if normalized[: len(expected)] != expected:
                raise ValueError("Katsayilar sayfasındaki sütunlar beklenen şablonla uyuşmuyor.")

            imported_rows: list[dict] = []
            for row_idx in range(2, ws_coeff.max_row + 1):
                values = [ws_coeff.cell(row=row_idx, column=i).value for i in range(1, 6)]
                if all(v in (None, "") for v in values):
                    continue
                bolge = normalize_key(values[0] or "")
                il = normalize_key(values[1] or "")
                ilce = normalize_key(values[2] or "")
                if not il or not ilce:
                    continue
                tapu = parse_decimal_input(values[3], 1.0)
                kadastro = parse_decimal_input(values[4], 1.0)
                imported_rows.append(
                    {
                        "sira": 20000 + len(imported_rows),
                        "bolge": bolge or il,
                        "il": il,
                        "ilce": ilce,
                        "tapu": float(tapu),
                        "kadastro": float(kadastro),
                    }
                )

            imported_rates = dict(self.control_rates)
            if "KontrollukUcretleri" in wb.sheetnames:
                ws_rates = wb["KontrollukUcretleri"]
                rate_header = [str(cell.value or "").strip() for cell in ws_rates[1]]
                normalized_rate = [canonical_header(v) for v in rate_header]
                expected_rate = [canonical_header(h) for h in RATE_EXCEL_HEADERS]
                if normalized_rate[: len(expected_rate)] != expected_rate:
                    raise ValueError("KontrollukUcretleri sayfasındaki sütunlar beklenen şablonla uyuşmuyor.")
                for row_idx in range(2, ws_rates.max_row + 1):
                    code = str(ws_rates.cell(row=row_idx, column=1).value or "").strip()
                    if not code or code not in CONTROL_RATE_DEFAULTS:
                        continue
                    amount = parse_decimal_input(ws_rates.cell(row=row_idx, column=3).value, CONTROL_RATE_DEFAULTS[code])
                    imported_rates[code] = float(amount)
        except ValueError as exc:
            messagebox.showerror("Hata", str(exc))
            return
        except Exception as exc:
            messagebox.showerror("Hata", f"Excel okunamadı: {exc}")
            return

        if not imported_rows:
            messagebox.showerror("Hata", "Excel dosyasında geçerli katsayı kaydı bulunamadı.")
            return

        custom_by_year = self.settings.setdefault("custom_rows_by_year", {})
        custom_by_year[str(self.selected_year)] = imported_rows
        self.control_rates = imported_rates
        self.settings["control_rates"] = self.control_rates
        self.save_settings()
        self.rebuild_data()
        self.refresh_control_iller()
        self.refresh_admin_ui()
        self.refresh_mahkeme_infaz_label()
        messagebox.showinfo(
            "Bilgi",
            (
                f"{self.selected_year} yılı için {len(imported_rows)} satır katsayı güncellendi.\n"
                f"Kontrollük tutarları: {len(self.control_rates)} kayıt güncel."
            ),
        )

    def on_admin_row_double_click(self, _event: object | None = None) -> None:
        selected = self.admin_tree.selection()
        if not selected:
            return
        vals = self.admin_tree.item(selected[0], "values")
        if len(vals) != 5:
            return
        row_data = {
            "bolge": str(vals[0]),
            "il": str(vals[1]),
            "ilce": str(vals[2]),
            "tapu": str(vals[3]),
            "kadastro": str(vals[4]),
        }
        self.open_admin_edit_dialog(row_data)

    def on_rate_row_double_click(self, _event: object | None = None) -> None:
        selected = self.rate_tree.selection()
        if not selected:
            return
        vals = self.rate_tree.item(selected[0], "values")
        if len(vals) != 3:
            return
        code = str(vals[0]).strip()
        label = str(vals[1]).strip()
        value = self.control_rates.get(code, CONTROL_RATE_DEFAULTS.get(code, 0.0))
        self.open_rate_edit_dialog(code, label, value)

    def open_admin_edit_dialog(self, row_data: dict[str, str]) -> None:
        dialog = tk.Toplevel(self)
        dialog.title("Katsayı Güncelle")
        dialog.geometry("520x250")
        dialog.transient(self)
        dialog.grab_set()

        frm = ttk.Frame(dialog, padding=12)
        frm.pack(fill="both", expand=True)
        vars_map = {
            "bolge": tk.StringVar(value=row_data.get("bolge", "")),
            "il": tk.StringVar(value=row_data.get("il", "")),
            "ilce": tk.StringVar(value=row_data.get("ilce", "")),
            "tapu": tk.StringVar(value=row_data.get("tapu", "")),
            "kadastro": tk.StringVar(value=row_data.get("kadastro", "")),
        }
        fields = [
            ("Bölge", "bolge"),
            ("İl", "il"),
            ("İlçe", "ilce"),
            ("Tapu YK", "tapu"),
            ("Kadastro YK", "kadastro"),
        ]
        for idx, (label, key) in enumerate(fields):
            ttk.Label(frm, text=label).grid(row=idx, column=0, sticky="w", padx=(0, 8), pady=6)
            ttk.Entry(frm, textvariable=vars_map[key], width=30).grid(row=idx, column=1, sticky="w", pady=6)

        def save_edit() -> None:
            il = normalize_key(vars_map["il"].get())
            ilce = normalize_key(vars_map["ilce"].get())
            bolge = normalize_key(vars_map["bolge"].get()) or il
            if not il or not ilce:
                messagebox.showerror("Hata", "İl ve ilçe boş olamaz.")
                return
            try:
                tapu = parse_decimal_input(vars_map["tapu"].get(), 1.0)
                kadastro = parse_decimal_input(vars_map["kadastro"].get(), 1.0)
            except ValueError:
                messagebox.showerror("Hata", "Tapu/Kadastro YK alanları sayısal olmalıdır.")
                return

            custom_by_year = self.settings.setdefault("custom_rows_by_year", {})
            custom = list(custom_by_year.get(str(self.selected_year), []))
            new_row = {
                "sira": 30000 + len(custom),
                "bolge": bolge,
                "il": il,
                "ilce": ilce,
                "tapu": float(tapu),
                "kadastro": float(kadastro),
            }
            custom[:] = [
                r
                for r in custom
                if not (normalize_key(str(r.get("il", ""))) == il and normalize_key(str(r.get("ilce", ""))) == ilce)
            ]
            custom.append(new_row)
            custom_by_year[str(self.selected_year)] = custom
            self.save_settings()
            self.rebuild_data()
            self.refresh_control_iller()
            self.refresh_admin_ui()
            dialog.destroy()
            messagebox.showinfo("Bilgi", f"{self.selected_year} yılı için {il} / {ilce} güncellendi.")

        buttons = ttk.Frame(frm)
        buttons.grid(row=len(fields), column=0, columnspan=2, sticky="w", pady=(12, 0))
        ttk.Button(buttons, text="Kaydet", style="Primary.TButton", command=save_edit).pack(side="left")
        ttk.Button(buttons, text="İptal", command=dialog.destroy).pack(side="left", padx=8)

    def open_rate_edit_dialog(self, code: str, label: str, value: float) -> None:
        dialog = tk.Toplevel(self)
        dialog.title("Kontrollük Tutarı Güncelle")
        dialog.geometry("560x220")
        dialog.transient(self)
        dialog.grab_set()

        frm = ttk.Frame(dialog, padding=12)
        frm.pack(fill="both", expand=True)
        ttk.Label(frm, text="Kod:", style="Muted.TLabel").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=6)
        ttk.Label(frm, text=code, style="Info.TLabel").grid(row=0, column=1, sticky="w", pady=6)
        ttk.Label(frm, text="Açıklama:", style="Muted.TLabel").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=6)
        ttk.Label(frm, text=label, wraplength=360, justify="left").grid(row=1, column=1, sticky="w", pady=6)
        amount_var = tk.StringVar(value=f"{value:.2f}".replace(".", ","))
        ttk.Label(frm, text="Tutar:", style="Muted.TLabel").grid(row=2, column=0, sticky="w", padx=(0, 8), pady=6)
        ttk.Entry(frm, textvariable=amount_var, width=24).grid(row=2, column=1, sticky="w", pady=6)

        def save_edit() -> None:
            try:
                amount = parse_decimal_input(amount_var.get(), value)
            except ValueError:
                messagebox.showerror("Hata", "Tutar alanı sayısal olmalıdır.")
                return
            if amount < 0:
                messagebox.showerror("Hata", "Tutar negatif olamaz.")
                return
            self.control_rates[code] = float(amount)
            self.settings["control_rates"] = self.control_rates
            self.save_settings()
            self.refresh_rate_ui()
            self.refresh_mahkeme_infaz_label()
            dialog.destroy()
            messagebox.showinfo("Bilgi", f"{code} tutarı güncellendi.")

        buttons = ttk.Frame(frm)
        buttons.grid(row=3, column=0, columnspan=2, sticky="w", pady=(12, 0))
        ttk.Button(buttons, text="Kaydet", style="Primary.TButton", command=save_edit).pack(side="left")
        ttk.Button(buttons, text="İptal", command=dialog.destroy).pack(side="left", padx=8)

    def build_about_tab(self) -> None:
        outer = ttk.Frame(self.tab_about)
        outer.pack(fill="both", expand=True)
        outer.grid_columnconfigure(0, weight=1)
        outer.grid_rowconfigure(0, weight=1)

        p = self.current_palette or {"bg": "#eef3fb", "panel": "#ffffff", "border": "#c8d6eb"}
        card = tk.Frame(
            outer,
            bg=p["panel"],
            highlightbackground=p["border"],
            highlightthickness=1,
            bd=0,
            padx=28,
            pady=24,
        )
        card.grid(row=0, column=0, sticky="nsew", padx=80, pady=36)
        self.about_card = card

        header = ttk.Frame(card)
        header.pack(fill="x", pady=(0, 18))
        ttk.Label(header, text="Hakkında", style="Title.TLabel").pack(anchor="w")
        ttk.Label(
            header,
            text=(
                "Lütfen sonuçları tarife cetveli ile test ederek işlemine devam ediniz. "
                "Harç hesaplamasından kaynaklı yasal sorumluluk kullanıcısına aittir."
            ),
            style="Muted.TLabel",
            wraplength=860,
            justify="left",
        ).pack(anchor="w", pady=(4, 0))

        info = ttk.Frame(card)
        info.pack(fill="x", pady=(0, 12))
        info.grid_columnconfigure(0, weight=1)
        info.grid_columnconfigure(1, weight=1)

        def add_column(parent: ttk.Frame, col: int, rows: list[tuple[str, str, str]]) -> None:
            column = ttk.Frame(parent)
            column.grid(row=0, column=col, sticky="nsew", padx=(0 if col == 0 else 12, 12 if col == 0 else 0))
            for icon, label, value in rows:
                line = ttk.Frame(column)
                line.pack(fill="x", pady=6)
                ttk.Label(line, text=icon, font=("Segoe UI Emoji", 14)).pack(side="left", padx=(0, 10))
                ttk.Label(line, text=f"{label}:", width=12, style="Muted.TLabel").pack(side="left")
                ttk.Label(line, text=value, style="Info.TLabel").pack(side="left")

        add_column(
            info,
            0,
            [
                ("👤", "Hazırlayan", "Fatih AKTAŞ"),
                ("🏢", "Kurum", "Elazığ Kadastro Müdürlüğü"),
                ("✉", "E-posta", "tk42980@tkgm.gov.tr"),
                ("☎", "Telefon", "-"),
            ],
        )
        add_column(
            info,
            1,
            [
                ("👤", "Hazırlayan", "Halil ALTIN"),
                ("🏢", "Kurum", "Muş Kadastro Müdürlüğü"),
                ("✉", "E-posta", "tk45410@tkgm.gov.tr"),
                ("☎", "Telefon", "84361002"),
            ],
        )


def main() -> None:
    app = KadastroDesktopApp()
    app.mainloop()


if __name__ == "__main__":
    main()
